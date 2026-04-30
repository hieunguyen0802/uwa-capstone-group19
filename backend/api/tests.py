import io

from django.contrib.auth.models import User
from rest_framework.test import APITestCase, APIClient
from rest_framework_simplejwt.tokens import RefreshToken
import openpyxl
from .models import Department, Staff, WorkloadReport


# ─── Shared test fixture ──────────────────────────────────────────────────────

class BaseTestCase(APITestCase):
    """
    Creates two departments and one staff member per role.
    All tests inherit from this so they don't repeat setup code.

    Java equivalent: a @BeforeEach method in a JUnit base class.
    """

    def setUp(self):
        # Two departments — used to verify HOD cannot cross department boundaries
        self.dept_csse = Department.objects.create(name='CSSE')
        self.dept_physics = Department.objects.create(name='Physics')

        # One staff per role
        self.academic = self._make_staff('academic1', 'ACADEMIC', self.dept_csse)
        self.hod_csse  = self._make_staff('hod_csse',  'HOD',      self.dept_csse)
        self.hod_phys  = self._make_staff('hod_phys',  'HOD',      self.dept_physics)
        self.ops       = self._make_staff('ops1',       'SCHOOL_OPS', self.dept_csse)
        self.hos       = self._make_staff('hos1',       'HOS',      self.dept_csse)

        # One PENDING report belonging to the CSSE academic
        self.report = WorkloadReport.objects.create(
            staff=self.academic,
            academic_year=2025,
            semester='S1',
            snapshot_fte=self.academic.fte,
            snapshot_department=self.dept_csse,
            status='PENDING',
        )

    def _make_staff(self, username, role, department):
        user = User.objects.create_user(username=username, password='testpass')
        # Use a hash suffix to guarantee uniqueness across all test methods
        import hashlib
        staff_number = hashlib.md5(username.encode()).hexdigest()[:8]
        return Staff.objects.create(
            user=user,
            staff_number=staff_number,
            role=role,
            department=department,
        )

    def _auth_client(self, staff):
        """Return an APIClient with a valid JWT for the given staff member."""
        token = RefreshToken.for_user(staff.user).access_token
        client = APIClient()
        client.credentials(HTTP_AUTHORIZATION=f'Bearer {token}')
        return client


# ─── Test: require_role decorator ────────────────────────────────────────────

class TestRequireRole(BaseTestCase):
    """
    Verifies that the @require_role decorator blocks wrong roles with 403.
    These are pure permission-boundary tests — no business logic involved.
    """

    def test_academic_cannot_access_supervisor_list(self):
        # ACADEMIC calling a HOD/OPS/HOS-only endpoint must get 403
        client = self._auth_client(self.academic)
        res = client.get('/api/supervisor/requests/')
        self.assertEqual(res.status_code, 403)

    def test_hod_can_access_supervisor_list(self):
        client = self._auth_client(self.hod_csse)
        res = client.get('/api/supervisor/requests/')
        self.assertEqual(res.status_code, 200)

    def test_unauthenticated_gets_401(self):
        # No token at all — DRF JWT middleware returns 401 before our decorator runs
        res = self.client.get('/api/supervisor/requests/')
        self.assertEqual(res.status_code, 401)

    def test_hod_cannot_access_academic_endpoint(self):
        # HOD calling an ACADEMIC-only endpoint must get 403
        client = self._auth_client(self.hod_csse)
        res = client.get('/api/workloads/my/')
        self.assertEqual(res.status_code, 403)

    def test_academic_can_access_own_workloads(self):
        client = self._auth_client(self.academic)
        res = client.get('/api/workloads/my/')
        self.assertEqual(res.status_code, 200)


# ─── Test: HOD data isolation ─────────────────────────────────────────────────

class TestHODDataIsolation(BaseTestCase):
    """
    Verifies that HOD can only see and act on reports in their own department.
    The report fixture belongs to dept_csse.
    """

    def test_hod_csse_sees_csse_report_in_list(self):
        client = self._auth_client(self.hod_csse)
        res = client.get('/api/supervisor/requests/')
        self.assertEqual(res.status_code, 200)
        all_ids = (
            [r['report_id'] for r in res.data['pending']] +
            [r['report_id'] for r in res.data['approved']] +
            [r['report_id'] for r in res.data['history']]
        )
        self.assertIn(str(self.report.report_id), all_ids)

    def test_hod_physics_cannot_see_csse_report(self):
        # Physics HOD's queryset is filtered to dept_physics — CSSE report is invisible
        client = self._auth_client(self.hod_phys)
        res = client.get('/api/supervisor/requests/')
        self.assertEqual(res.status_code, 200)
        all_ids = (
            [r['report_id'] for r in res.data['pending']] +
            [r['report_id'] for r in res.data['approved']] +
            [r['report_id'] for r in res.data['history']]
        )
        self.assertNotIn(str(self.report.report_id), all_ids)

    def test_hod_physics_cannot_approve_csse_report(self):
        # get_workload_queryset filters by dept, so the report is not found → 404
        client = self._auth_client(self.hod_phys)
        res = client.post(f'/api/supervisor/approve/{self.report.report_id}/')
        self.assertEqual(res.status_code, 404)

    def test_hod_csse_can_approve_csse_report(self):
        client = self._auth_client(self.hod_csse)
        res = client.post(f'/api/supervisor/approve/{self.report.report_id}/')
        self.assertEqual(res.status_code, 200)
        self.report.refresh_from_db()
        self.assertEqual(self.report.status, 'APPROVED')


# ─── Test: approval state machine ────────────────────────────────────────────

class TestApprovalStateMachine(BaseTestCase):
    """
    Verifies that the status machine rejects illegal transitions.
    """

    def test_cannot_approve_already_approved_report(self):
        # Approve once — should succeed
        client = self._auth_client(self.hod_csse)
        client.post(f'/api/supervisor/approve/{self.report.report_id}/')

        # Approve again — should get 409 Conflict
        res = client.post(f'/api/supervisor/approve/{self.report.report_id}/')
        self.assertEqual(res.status_code, 409)

    def test_reject_requires_comment(self):
        client = self._auth_client(self.hod_csse)
        res = client.post(
            f'/api/supervisor/reject/{self.report.report_id}/',
            data={},  # no comment
            format='json'
        )
        self.assertEqual(res.status_code, 422)

    def test_reject_with_comment_succeeds(self):
        client = self._auth_client(self.hod_csse)
        res = client.post(
            f'/api/supervisor/reject/{self.report.report_id}/',
            data={'comment': 'Teaching hours look incorrect.'},
            format='json'
        )
        self.assertEqual(res.status_code, 200)
        self.report.refresh_from_db()
        self.assertEqual(self.report.status, 'REJECTED')

    def test_academic_cannot_submit_twice(self):
        # First query submission
        client = self._auth_client(self.academic)
        client.post(
            '/api/queries/',
            data={'workload_report_id': str(self.report.report_id), 'comment': 'I disagree with this.'},
            format='json'
        )
        # Second submission on same report — already has a COMMENT log, should fail
        res = client.post(
            '/api/queries/',
            data={'workload_report_id': str(self.report.report_id), 'comment': 'Trying again.'},
            format='json'
        )
        self.assertEqual(res.status_code, 409)


# ─── Test: Academic workload view + query submission (#3) ─────────────────────

class TestAcademicWorkloadAndQuery(BaseTestCase):
    """
    Verifies GET /api/workloads/my/ and POST /api/queries/ behaviour.
    """

    def test_academic_sees_own_report(self):
        client = self._auth_client(self.academic)
        res = client.get('/api/workloads/my/')
        self.assertEqual(res.status_code, 200)
        ids = [r['report_id'] for r in res.data]
        self.assertIn(str(self.report.report_id), ids)

    def test_academic_does_not_see_other_report(self):
        # Create a second academic in the same dept and verify they cannot see each other's reports
        other = self._make_staff('academic2', 'ACADEMIC', self.dept_csse)
        client = self._auth_client(other)
        res = client.get('/api/workloads/my/')
        self.assertEqual(res.status_code, 200)
        ids = [r['report_id'] for r in res.data]
        self.assertNotIn(str(self.report.report_id), ids)

    def test_submit_query_creates_audit_log(self):
        from api.models import AuditLog
        client = self._auth_client(self.academic)
        res = client.post(
            '/api/queries/',
            data={'workload_report_id': str(self.report.report_id), 'comment': 'Hours look wrong.'},
            format='json'
        )
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data['report_id'], str(self.report.report_id))
        self.assertTrue(
            AuditLog.objects.filter(report=self.report, action_type='COMMENT').exists()
        )

    def test_submit_query_missing_comment_returns_400(self):
        client = self._auth_client(self.academic)
        res = client.post(
            '/api/queries/',
            data={'workload_report_id': str(self.report.report_id)},
            format='json'
        )
        self.assertEqual(res.status_code, 400)

    def test_submit_query_missing_report_id_returns_400(self):
        client = self._auth_client(self.academic)
        res = client.post(
            '/api/queries/',
            data={'comment': 'Something is off.'},
            format='json'
        )
        self.assertEqual(res.status_code, 400)

    def test_hod_cannot_submit_query(self):
        # POST /api/queries/ is ACADEMIC-only
        client = self._auth_client(self.hod_csse)
        res = client.post(
            '/api/queries/',
            data={'workload_report_id': str(self.report.report_id), 'comment': 'test'},
            format='json'
        )
        self.assertEqual(res.status_code, 403)

    def test_academic_cannot_query_other_academic_report(self):
        # Academic2 tries to query Academic1's report — should get 404
        other = self._make_staff('academic3', 'ACADEMIC', self.dept_csse)
        client = self._auth_client(other)
        res = client.post(
            '/api/queries/',
            data={'workload_report_id': str(self.report.report_id), 'comment': 'Not mine.'},
            format='json'
        )
        self.assertEqual(res.status_code, 404)


# ─── Test: HOD query approval (#4) ───────────────────────────────────────────

class TestHODQueryApproval(BaseTestCase):
    """
    Verifies GET /api/queries/pending/ and POST /api/queries/<id>/approve|reject/.

    setUp creates a COMMENT log on self.report so it appears in the pending list.
    Tests cover: data isolation, state machine, validation, role boundaries.
    """

    def setUp(self):
        super().setUp()
        # Simulate Academic submitting a query so the report appears in HOD's pending list
        from api.models import AuditLog
        AuditLog.objects.create(
            report=self.report,
            action_by=self.academic,
            action_type='COMMENT',
            comment='I disagree with the teaching hours.',
        )

    # ── GET /api/queries/pending/ ─────────────────────────────────────────────

    def test_hod_sees_queried_report_in_pending_list(self):
        # Report has a COMMENT log → must appear in HOD's pending list
        client = self._auth_client(self.hod_csse)
        res = client.get('/api/queries/pending/')
        self.assertEqual(res.status_code, 200)
        ids = [r['report_id'] for r in res.data]
        self.assertIn(str(self.report.report_id), ids)

    def test_hod_sees_report_without_query_too(self):
        # A PENDING report with no COMMENT log must still appear — HOD reviews all PENDING reports
        other_report = WorkloadReport.objects.create(
            staff=self.academic,
            academic_year=2025,
            semester='S2',
            snapshot_fte=self.academic.fte,
            snapshot_department=self.dept_csse,
            status='PENDING',
        )
        client = self._auth_client(self.hod_csse)
        res = client.get('/api/queries/pending/')
        self.assertEqual(res.status_code, 200)
        ids = [r['report_id'] for r in res.data]
        self.assertIn(str(other_report.report_id), ids)

    def test_hod_physics_cannot_see_csse_query(self):
        # Physics HOD's queryset is scoped to dept_physics — CSSE report is invisible
        client = self._auth_client(self.hod_phys)
        res = client.get('/api/queries/pending/')
        self.assertEqual(res.status_code, 200)
        ids = [r['report_id'] for r in res.data]
        self.assertNotIn(str(self.report.report_id), ids)

    def test_academic_cannot_access_pending_queries(self):
        client = self._auth_client(self.academic)
        res = client.get('/api/queries/pending/')
        self.assertEqual(res.status_code, 403)

    # ── POST /api/queries/<id>/approve/ ──────────────────────────────────────

    def test_hod_can_approve_query(self):
        from api.models import AuditLog
        client = self._auth_client(self.hod_csse)
        res = client.post(
            f'/api/queries/{self.report.report_id}/approve/',
            data={'reason': 'Hours look correct after review.'},
            format='json',
        )
        self.assertEqual(res.status_code, 200)
        self.report.refresh_from_db()
        self.assertEqual(self.report.status, 'APPROVED')
        self.assertTrue(
            AuditLog.objects.filter(report=self.report, action_type='APPROVE').exists()
        )

    def test_hod_can_approve_without_reason(self):
        # reason is optional for approve
        client = self._auth_client(self.hod_csse)
        res = client.post(
            f'/api/queries/{self.report.report_id}/approve/',
            data={},
            format='json',
        )
        self.assertEqual(res.status_code, 200)

    def test_hod_physics_cannot_approve_csse_query(self):
        # get_workload_queryset scopes to dept_physics → report not found → 404
        client = self._auth_client(self.hod_phys)
        res = client.post(
            f'/api/queries/{self.report.report_id}/approve/',
            data={'reason': 'Looks fine.'},
            format='json',
        )
        self.assertEqual(res.status_code, 404)

    def test_cannot_approve_already_approved_query(self):
        client = self._auth_client(self.hod_csse)
        client.post(f'/api/queries/{self.report.report_id}/approve/', data={}, format='json')
        res = client.post(f'/api/queries/{self.report.report_id}/approve/', data={}, format='json')
        self.assertEqual(res.status_code, 409)

    # ── POST /api/queries/<id>/reject/ ───────────────────────────────────────

    def test_hod_can_reject_query_with_reason(self):
        from api.models import AuditLog
        client = self._auth_client(self.hod_csse)
        res = client.post(
            f'/api/queries/{self.report.report_id}/reject/',
            data={'reason': 'Teaching hours are within the agreed allocation.'},
            format='json',
        )
        self.assertEqual(res.status_code, 200)
        self.report.refresh_from_db()
        self.assertEqual(self.report.status, 'REJECTED')
        self.assertTrue(
            AuditLog.objects.filter(report=self.report, action_type='REJECT').exists()
        )

    def test_hod_reject_without_reason_returns_422(self):
        client = self._auth_client(self.hod_csse)
        res = client.post(
            f'/api/queries/{self.report.report_id}/reject/',
            data={},
            format='json',
        )
        self.assertEqual(res.status_code, 422)

    def test_cannot_reject_already_rejected_query(self):
        client = self._auth_client(self.hod_csse)
        client.post(
            f'/api/queries/{self.report.report_id}/reject/',
            data={'reason': 'First rejection.'},
            format='json',
        )
        res = client.post(
            f'/api/queries/{self.report.report_id}/reject/',
            data={'reason': 'Second rejection.'},
            format='json',
        )
        self.assertEqual(res.status_code, 409)

    def test_academic_cannot_call_approve_endpoint(self):
        client = self._auth_client(self.academic)
        res = client.post(
            f'/api/queries/{self.report.report_id}/approve/',
            data={},
            format='json',
        )
        self.assertEqual(res.status_code, 403)


# ─── Test: HOS approval of HOD workload (#5) ─────────────────────────────────

class TestHOSApproval(BaseTestCase):
    """
    Verifies GET /api/workloads/hod-summary/ and
    POST /api/workloads/<id>/hos-approve|hos-reject/.

    setUp creates a WorkloadReport owned by hod_csse so HOS has something to review.
    """

    def setUp(self):
        super().setUp()
        # HOD's own workload report — the subject of HOS approval
        self.hod_report = WorkloadReport.objects.create(
            staff=self.hod_csse,
            academic_year=2025,
            semester='S1',
            snapshot_fte=self.hod_csse.fte,
            snapshot_department=self.dept_csse,
            status='PENDING',
        )

    # ── GET /api/workloads/hod-summary/ ──────────────────────────────────────

    def test_hos_sees_hod_report_in_summary(self):
        client = self._auth_client(self.hos)
        res = client.get('/api/workloads/hod-summary/')
        self.assertEqual(res.status_code, 200)
        ids = [r['report_id'] for r in res.data]
        self.assertIn(str(self.hod_report.report_id), ids)

    def test_academic_report_not_in_hod_summary(self):
        # Academic's report must NOT appear — only HOD reports are listed
        client = self._auth_client(self.hos)
        res = client.get('/api/workloads/hod-summary/')
        ids = [r['report_id'] for r in res.data]
        self.assertNotIn(str(self.report.report_id), ids)

    def test_non_hos_cannot_access_hod_summary(self):
        for staff in [self.academic, self.hod_csse, self.ops]:
            client = self._auth_client(staff)
            res = client.get('/api/workloads/hod-summary/')
            self.assertEqual(res.status_code, 403)

    # ── POST /api/workloads/<id>/hos-approve/ ────────────────────────────────

    def test_hos_can_approve_hod_report(self):
        from api.models import AuditLog
        client = self._auth_client(self.hos)
        res = client.post(
            f'/api/workloads/{self.hod_report.report_id}/hos-approve/',
            data={'reason': 'Workload allocation looks correct.'},
            format='json',
        )
        self.assertEqual(res.status_code, 200)
        self.hod_report.refresh_from_db()
        self.assertEqual(self.hod_report.status, 'APPROVED')
        self.assertTrue(
            AuditLog.objects.filter(report=self.hod_report, action_type='APPROVE').exists()
        )

    def test_hos_can_approve_without_reason(self):
        client = self._auth_client(self.hos)
        res = client.post(
            f'/api/workloads/{self.hod_report.report_id}/hos-approve/',
            data={},
            format='json',
        )
        self.assertEqual(res.status_code, 200)

    def test_cannot_approve_already_approved_hod_report(self):
        self.hod_report.status = 'APPROVED'
        self.hod_report.save()
        client = self._auth_client(self.hos)
        res = client.post(
            f'/api/workloads/{self.hod_report.report_id}/hos-approve/',
            data={},
            format='json',
        )
        self.assertEqual(res.status_code, 409)

    def test_non_hos_cannot_approve_hod_report(self):
        for staff in [self.academic, self.hod_csse, self.ops]:
            client = self._auth_client(staff)
            res = client.post(
                f'/api/workloads/{self.hod_report.report_id}/hos-approve/',
                data={},
                format='json',
            )
            self.assertEqual(res.status_code, 403)

    def test_hos_cannot_approve_academic_report(self):
        # self.report belongs to an ACADEMIC — must return 404 (not a HOD report)
        client = self._auth_client(self.hos)
        res = client.post(
            f'/api/workloads/{self.report.report_id}/hos-approve/',
            data={},
            format='json',
        )
        self.assertEqual(res.status_code, 404)

    # ── POST /api/workloads/<id>/hos-reject/ ─────────────────────────────────

    def test_hos_can_reject_hod_report_with_reason(self):
        from api.models import AuditLog
        client = self._auth_client(self.hos)
        res = client.post(
            f'/api/workloads/{self.hod_report.report_id}/hos-reject/',
            data={'reason': 'Teaching allocation exceeds agreed FTE.'},
            format='json',
        )
        self.assertEqual(res.status_code, 200)
        self.hod_report.refresh_from_db()
        self.assertEqual(self.hod_report.status, 'REJECTED')
        self.assertTrue(
            AuditLog.objects.filter(report=self.hod_report, action_type='REJECT').exists()
        )

    def test_hos_reject_without_reason_returns_422(self):
        client = self._auth_client(self.hos)
        res = client.post(
            f'/api/workloads/{self.hod_report.report_id}/hos-reject/',
            data={},
            format='json',
        )
        self.assertEqual(res.status_code, 422)

    def test_cannot_reject_already_rejected_hod_report(self):
        self.hod_report.status = 'REJECTED'
        self.hod_report.save()
        client = self._auth_client(self.hos)
        res = client.post(
            f'/api/workloads/{self.hod_report.report_id}/hos-reject/',
            data={'reason': 'Second rejection attempt.'},
            format='json',
        )
        self.assertEqual(res.status_code, 409)


# ─── Test: Excel import (#6) ─────────────────────────────────────────────────

class TestExcelImport(BaseTestCase):
    """Verifies POST /api/ops/import/ for SCHOOL_OPS-only Excel imports."""

    def _make_excel_bytes(self, *, semester='S1', academic_year=2025, department='CSSE',
                          staff_number=None, staff_type='Academic staff'):
        wb = openpyxl.Workbook()
        ws = wb.active

        # File-level context expected by import service
        ws.cell(row=1, column=2, value=semester)        # B1
        ws.cell(row=1, column=3, value=academic_year)   # C1
        ws.cell(row=1, column=4, value=department)      # D1

        headers = [
            'Staff Member ID', 'Staff Name', 'Staff Number', 'FTE', 'Function',
            'Target Band', 'Target Teaching %', 'Unit Code', 'Unit Enrolment', 'Staff Type',
            'Teaching Hrs', 'Teaching WL Pts', 'Unit Coord Hrs', 'Unit Coord WL Pts',
            'Teaching Activity Hrs', 'Teaching Activity WL Pts', 'Unit Supervision Hrs',
            'Unit Supervision WL Pts', 'New Unit Dev Hrs', 'New Unit Dev WL Pts',
            'Total Teaching WL Pts', 'FT Students', 'FT Proportion', 'PT Students',
            'PT Proportion', 'HDR Total Hrs', 'HDR WL Pts', 'Self-Directed Svc Pts',
            'Assigned Roles Total Pts', 'Role 1 Name', 'Role 1 Points', 'Role 2 Name',
            'Role 2 Points', 'Role 3 Name'
        ]
        for idx, h in enumerate(headers, 1):
            ws.cell(row=4, column=idx, value=h)

        row_staff_number = staff_number or self.academic.staff_number
        row = [
            f'Test User, {row_staff_number}', 'Test, User', row_staff_number,
            1.0, 'T & R', 'Balanced Teaching & Research', 50,
            'CITS5206', 100, staff_type,
            172.5, 10, 0, 0, 0, 0, 0, 0, 0, 0,
            10, 2, 0.8, 0, 0, 86.25, 5, 10, 4,
            'Role A', 2.0, '', 0, ''
        ]
        for idx, v in enumerate(row, 1):
            ws.cell(row=5, column=idx, value=v)

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer

    def test_ops_can_import_excel(self):
        from api.models import AuditLog, WorkloadItem

        client = self._auth_client(self.ops)
        excel = self._make_excel_bytes(semester='S2')
        excel.name = 'workload_import.xlsx'

        res = client.post('/api/ops/import/', {'file': excel}, format='multipart')
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data['created'], 1)
        self.assertEqual(res.data['updated'], 0)

        report = WorkloadReport.objects.filter(staff=self.academic, semester='S2', academic_year=2025).latest('created_at')
        self.assertEqual(report.snapshot_department.name, 'CSSE')
        self.assertTrue(WorkloadItem.objects.filter(report=report).exists())
        self.assertTrue(AuditLog.objects.filter(report=report, action_type='IMPORTED').exists())

    def test_non_ops_cannot_import_excel(self):
        for staff in [self.academic, self.hod_csse, self.hos]:
            client = self._auth_client(staff)
            excel = self._make_excel_bytes()
            excel.name = 'workload_import.xlsx'
            res = client.post('/api/ops/import/', {'file': excel}, format='multipart')
            self.assertEqual(res.status_code, 403)

    def test_import_fails_when_header_fields_missing(self):
        client = self._auth_client(self.ops)
        excel = self._make_excel_bytes(semester='', academic_year='', department='')
        excel.name = 'workload_import.xlsx'

        res = client.post('/api/ops/import/', {'file': excel}, format='multipart')
        self.assertEqual(res.status_code, 400)
        self.assertEqual(res.data['code'], 'VALIDATION_ERROR')
        self.assertIn('Missing required header fields', res.data['message'])

    def test_import_skips_unknown_staff_and_reports_error(self):
        client = self._auth_client(self.ops)
        excel = self._make_excel_bytes(staff_number='UNKNOWN999')
        excel.name = 'workload_import.xlsx'

        res = client.post('/api/ops/import/', {'file': excel}, format='multipart')
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data['created'], 0)
        self.assertEqual(res.data['skipped'], 1)
        self.assertTrue(len(res.data['errors']) > 0)
        self.assertIn('UNKNOWN999', res.data['errors'][0])

    def test_import_skips_casual_staff(self):
        client = self._auth_client(self.ops)
        excel = self._make_excel_bytes(staff_type='Paid casual staff')
        excel.name = 'workload_import.xlsx'

        res = client.post('/api/ops/import/', {'file': excel}, format='multipart')
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data['created'], 0)
        self.assertEqual(res.data['skipped'], 1)

    def test_reimport_supersedes_pending_report(self):
        from api.models import WorkloadReport, AuditLog

        client = self._auth_client(self.ops)
        # First import — creates a PENDING report
        excel = self._make_excel_bytes(semester='S1', academic_year=2025)
        excel.name = 'workload_import.xlsx'
        res = client.post('/api/ops/import/', {'file': excel}, format='multipart')
        self.assertEqual(res.status_code, 201)
        self.assertEqual(res.data['created'], 1)

        # Second import of same staff/semester/year — should supersede
        excel2 = self._make_excel_bytes(semester='S1', academic_year=2025)
        excel2.name = 'workload_import.xlsx'
        res2 = client.post('/api/ops/import/', {'file': excel2}, format='multipart')
        self.assertEqual(res2.status_code, 201)
        self.assertEqual(res2.data['updated'], 1)

        # Old report must be marked non-current
        old = WorkloadReport.objects.filter(
            staff=self.academic, semester='S1', academic_year=2025, is_current=False
        )
        self.assertTrue(old.exists())
        # New report must be current
        new = WorkloadReport.objects.get(
            staff=self.academic, semester='S1', academic_year=2025, is_current=True
        )
        self.assertTrue(
            AuditLog.objects.filter(report=new, action_type='MODIFIED_BY_REIMPORT').exists()
        )

    def test_import_rejects_non_xlsx_file(self):
        client = self._auth_client(self.ops)
        fake_csv = io.BytesIO(b'col1,col2\nval1,val2')
        fake_csv.name = 'workload.csv'

        res = client.post('/api/ops/import/', {'file': fake_csv}, format='multipart')
        self.assertIn(res.status_code, [400, 422])
