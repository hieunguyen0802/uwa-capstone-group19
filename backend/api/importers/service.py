import uuid
from decimal import Decimal
from typing import Optional

import openpyxl

from api.importers.dto import WorkloadRowDTO
from api.models import AuditLog, Department, Staff, WorkloadItem, WorkloadReport

WL_HOURS_PER_POINT = Decimal('17.25')

# Column indices in the template (0-based, row 4 is the header row)
COL_STAFF_ID = 0
COL_STAFF_NAME = 1
COL_STAFF_NUMBER = 2
COL_FTE = 3
COL_FUNCTION = 4
COL_TARGET_BAND = 5
COL_TARGET_TEACHING_PCT = 6
COL_UNIT_CODE = 7
COL_UNIT_ENROLMENT = 8
COL_STAFF_TYPE = 9
COL_TEACHING_HRS = 10
COL_UNIT_COORD_HRS = 12
COL_TEACHING_ACTIVITY_HRS = 14
COL_UNIT_SUPERVISION_HRS = 16
COL_NEW_UNIT_DEV_HRS = 18
COL_HDR_TOTAL_HRS = 25
COL_ASSIGNED_ROLES_TOTAL_PTS = 28

# Header cells for file-level context (row 1, columns B/C/D)
# Daniela writes: Semester=S1, Academic Year=2025, Department=CSSE in these cells.
# If these are absent, the import will fail with a 400 error.
HEADER_ROW_SEMESTER = (1, 2)       # row 1, col B
HEADER_ROW_ACADEMIC_YEAR = (1, 3)  # row 1, col C
HEADER_ROW_DEPARTMENT = (1, 4)     # row 1, col D

DATA_START_ROW = 5  # data rows begin at row 5 (rows 1-4 are header/instructions)


def _float(val, default=0.0) -> float:
    try:
        return float(val) if val is not None else default
    except (TypeError, ValueError):
        return default


def _str(val) -> str:
    return str(val).strip() if val is not None else ''


def parse_excel(file_obj) -> tuple[list[WorkloadRowDTO], dict]:
    """
    Parse the uploaded Excel file into a list of WorkloadRowDTOs.
    Returns (rows, header_meta) where header_meta contains semester/year/department.
    Raises ValueError with a descriptive message if the file is malformed.
    """
    wb = openpyxl.load_workbook(file_obj, data_only=True)
    ws = wb.active

    # Read file-level context from header cells
    header_semester = _str(ws.cell(*HEADER_ROW_SEMESTER).value)
    header_year_raw = ws.cell(*HEADER_ROW_ACADEMIC_YEAR).value
    header_department = _str(ws.cell(*HEADER_ROW_DEPARTMENT).value)

    try:
        header_year = int(header_year_raw) if header_year_raw else 0
    except (TypeError, ValueError):
        header_year = 0

    rows = []
    for row in ws.iter_rows(min_row=DATA_START_ROW, values_only=True):
        staff_number = _str(row[COL_STAFF_NUMBER] if len(row) > COL_STAFF_NUMBER else None)
        if not staff_number:
            continue  # skip blank rows

        # Per-row context overrides file header if present
        semester = header_semester
        academic_year = header_year
        department = header_department

        rows.append(WorkloadRowDTO(
            staff_number=staff_number,
            staff_name=_str(row[COL_STAFF_NAME] if len(row) > COL_STAFF_NAME else None),
            fte=_float(row[COL_FTE] if len(row) > COL_FTE else None, 1.0),
            function=_str(row[COL_FUNCTION] if len(row) > COL_FUNCTION else None),
            target_band=_str(row[COL_TARGET_BAND] if len(row) > COL_TARGET_BAND else None),
            target_teaching_pct=_float(row[COL_TARGET_TEACHING_PCT] if len(row) > COL_TARGET_TEACHING_PCT else None),
            staff_type=_str(row[COL_STAFF_TYPE] if len(row) > COL_STAFF_TYPE else None),
            unit_code=_str(row[COL_UNIT_CODE] if len(row) > COL_UNIT_CODE else None) or None,
            unit_enrolment=int(_float(row[COL_UNIT_ENROLMENT] if len(row) > COL_UNIT_ENROLMENT else None)) or None,
            teaching_hrs=_float(row[COL_TEACHING_HRS] if len(row) > COL_TEACHING_HRS else None),
            unit_coord_hrs=_float(row[COL_UNIT_COORD_HRS] if len(row) > COL_UNIT_COORD_HRS else None),
            teaching_activity_hrs=_float(row[COL_TEACHING_ACTIVITY_HRS] if len(row) > COL_TEACHING_ACTIVITY_HRS else None),
            unit_supervision_hrs=_float(row[COL_UNIT_SUPERVISION_HRS] if len(row) > COL_UNIT_SUPERVISION_HRS else None),
            new_unit_dev_hrs=_float(row[COL_NEW_UNIT_DEV_HRS] if len(row) > COL_NEW_UNIT_DEV_HRS else None),
            hdr_total_hrs=_float(row[COL_HDR_TOTAL_HRS] if len(row) > COL_HDR_TOTAL_HRS else None),
            assigned_roles_total_pts=_float(row[COL_ASSIGNED_ROLES_TOTAL_PTS] if len(row) > COL_ASSIGNED_ROLES_TOTAL_PTS else None),
            semester=semester,
            academic_year=academic_year,
            department=department,
        ))

    header_meta = {
        'semester': header_semester,
        'academic_year': header_year,
        'department': header_department,
    }
    return rows, header_meta


def _detect_anomaly(fte: float, teaching_pts: Decimal, hdr_pts: Decimal,
                    role_pts: Decimal, service_pts: Decimal, target_band: str) -> Optional[bool]:
    """
    Implements the Nidhi PDF anomaly detection algorithm exactly.
    Returns True if anomaly detected, False if not, None if insufficient data.
    """
    research_pts = Decimal(str(fte)) * 100 - (teaching_pts + role_pts + service_pts + hdr_pts)
    total = teaching_pts + research_pts
    if total == 0:
        return None

    calc_tr = round(float(teaching_pts / total), 2)

    if calc_tr <= 0.20:
        calc_band = 'Research Focused'
    elif calc_tr <= 0.79:
        calc_band = 'Balanced Teaching & Research'
    else:
        calc_band = 'Teaching Focused'

    return calc_band != target_band


def _hrs_to_pts(hours: float) -> Decimal:
    return (Decimal(str(hours)) / WL_HOURS_PER_POINT).quantize(Decimal('0.01'))


def import_workload_excel(file_obj, uploaded_by) -> dict:
    """
    Parse the Excel file and persist WorkloadReports + WorkloadItems.
    Returns a summary dict: {batch_id, created, updated, skipped, errors}.

    Raises ValueError for missing header fields (semester/year/department).
    """
    rows, header_meta = parse_excel(file_obj)

    # Validate required header fields
    missing = [k for k, v in header_meta.items() if not v]
    if missing:
        raise ValueError(
            f"Missing required header fields: {', '.join(missing)}. "
            "Please add Semester (cell B1), Academic Year (cell C1), "
            "and Department (cell D1) to the file header."
        )

    batch_id = uuid.uuid4()
    created = updated = skipped = 0
    errors = []

    for dto in rows:
        # Skip casual staff — they don't enter the approval workflow
        if 'casual' in dto.staff_type.lower():
            skipped += 1
            continue

        try:
            staff = Staff.objects.select_related('department').get(
                staff_number=dto.staff_number
            )
        except Staff.DoesNotExist:
            errors.append(f"Staff {dto.staff_number} not found — row skipped.")
            skipped += 1
            continue

        try:
            department = Department.objects.get(name__iexact=dto.department)
        except Department.DoesNotExist:
            errors.append(f"Department '{dto.department}' not found — row for {dto.staff_number} skipped.")
            skipped += 1
            continue

        # Calculate WL points for anomaly detection
        teaching_pts = _hrs_to_pts(
            dto.teaching_hrs + dto.unit_coord_hrs + dto.teaching_activity_hrs
            + dto.unit_supervision_hrs + dto.new_unit_dev_hrs
        )
        hdr_pts = _hrs_to_pts(dto.hdr_total_hrs)
        role_pts = Decimal(str(dto.assigned_roles_total_pts))
        service_pts = Decimal(str(dto.fte)) * 10  # Self-Directed Service = FTE × 10

        is_anomaly = _detect_anomaly(
            dto.fte, teaching_pts, hdr_pts, role_pts, service_pts, dto.target_band
        )

        # Check for existing record with same natural key
        existing = WorkloadReport.objects.filter(
            staff=staff,
            semester=dto.semester,
            academic_year=dto.academic_year,
            is_current=True,
        ).first()

        if existing:
            if existing.status == 'PENDING':
                # Supersede the old PENDING record
                new_report = _create_report(
                    staff, department, dto, batch_id, is_anomaly
                )
                existing.is_current = False
                existing.superseded_by = new_report
                existing.save()
                AuditLog.objects.create(
                    report=new_report,
                    action_by=uploaded_by,
                    action_type='MODIFIED_BY_REIMPORT',
                    comment=f"Superseded report {existing.report_id}",
                )
                _create_items(new_report, dto)
                updated += 1
            else:
                # Terminal state — create a correction version, flag for OPS review
                new_report = _create_report(
                    staff, department, dto, batch_id, is_anomaly
                )
                AuditLog.objects.create(
                    report=new_report,
                    action_by=uploaded_by,
                    action_type='MODIFIED_BY_REIMPORT',
                    comment=f"Correction of terminal record {existing.report_id} — requires OPS review.",
                )
                _create_items(new_report, dto)
                updated += 1
        else:
            new_report = _create_report(staff, department, dto, batch_id, is_anomaly)
            AuditLog.objects.create(
                report=new_report,
                action_by=uploaded_by,
                action_type='IMPORTED',
                comment=None,
            )
            _create_items(new_report, dto)
            created += 1

    return {
        'batch_id': str(batch_id),
        'created': created,
        'updated': updated,
        'skipped': skipped,
        'errors': errors,
    }


def _create_report(staff, department, dto: WorkloadRowDTO, batch_id, is_anomaly) -> WorkloadReport:
    return WorkloadReport.objects.create(
        staff=staff,
        academic_year=dto.academic_year,
        semester=dto.semester,
        snapshot_fte=Decimal(str(dto.fte)),
        snapshot_department=department,
        is_anomaly=bool(is_anomaly) if is_anomaly is not None else False,
        import_batch_id=batch_id,
        is_current=True,
        status='PENDING',
    )


def _create_items(report: WorkloadReport, dto: WorkloadRowDTO):
    items = []
    teaching_hrs = (
        dto.teaching_hrs + dto.unit_coord_hrs + dto.teaching_activity_hrs
        + dto.unit_supervision_hrs + dto.new_unit_dev_hrs
    )
    if teaching_hrs > 0:
        items.append(WorkloadItem(
            report=report,
            category='TEACHING',
            unit_code=dto.unit_code,
            allocated_hours=Decimal(str(teaching_hrs)),
        ))
    if dto.hdr_total_hrs > 0:
        items.append(WorkloadItem(
            report=report,
            category='HDR_SUPERVISION',
            allocated_hours=Decimal(str(dto.hdr_total_hrs)),
        ))
    if dto.assigned_roles_total_pts > 0:
        # Assigned roles are stored as points × 17.25 to convert back to hours
        role_hrs = float(dto.assigned_roles_total_pts) * float(WL_HOURS_PER_POINT)
        items.append(WorkloadItem(
            report=report,
            category='ASSIGNED_ROLE',
            allocated_hours=Decimal(str(role_hrs)).quantize(Decimal('0.01')),
        ))
    # Self-Directed Service is always auto-calculated
    service_hrs = float(dto.fte) * 10 * float(WL_HOURS_PER_POINT)
    items.append(WorkloadItem(
        report=report,
        category='SERVICE',
        allocated_hours=Decimal(str(service_hrs)).quantize(Decimal('0.01')),
    ))
    WorkloadItem.objects.bulk_create(items)
