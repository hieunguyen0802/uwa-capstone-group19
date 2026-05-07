"""
School Operations endpoints.

Authenticated roles: SCHOOL_OPS, HOS (school-wide visibility).

Routes are registered under both /api/school-operations/* (new contract) and
/api/admin/* (legacy alias kept for backward compatibility).
Comments are English-only per project guideline.
"""

import io
import re
import uuid
from decimal import Decimal
from pathlib import Path

from django.conf import settings
from django.core.cache import cache
from django.core.paginator import Paginator
from django.db import models, transaction
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from rest_framework import status as http_status
from rest_framework.decorators import api_view, permission_classes, throttle_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.throttling import UserRateThrottle

from api.decorators import require_role
from api.models import (
    AuditLog,
    Department,
    Staff,
    StaffRoleAssignment,
    WorkloadDistributionJob,
    WorkloadItem,
    WorkloadReport,
)
from api.services.workload_service import (
    STALE_REPORT_ERROR,
    get_workload_queryset,
    _filter_reports_by_range,
    _parse_year_range,
    persist_report_anomaly,
    evaluate_mvp_anomaly,
)
from api.view.supervisor_views import (
    _get_request_reason,
    _get_supervisor_note,
    _parse_breakdown_data,
    _to_decimal_hours,
)

ADMIN_ROLES = ('SCHOOL_OPS', 'HOS')
MAX_EXCEL_UPLOAD_BYTES = 5 * 1024 * 1024
EXPORT_MEDIA_SUBDIR = 'exports'
TEMPLATE_MEDIA_SUBDIR = 'templates'


class AdminImportThrottle(UserRateThrottle):
    rate = '30/hour'


class AdminExportThrottle(UserRateThrottle):
    rate = '60/hour'


def _ensure_media_subdir(segment: str) -> Path:
    base = Path(settings.MEDIA_ROOT) / segment
    base.mkdir(parents=True, exist_ok=True)
    return base


def _admin_reports_qs(staff):
    """
    Ops/HoS can see every current report (no Hod visibility gate).
    This matches school-wide dashboards while keeping academic/Hod isolation intact.
    """
    return get_workload_queryset(staff).filter(is_current=True)


def _parse_semester_filter(request):
    """Accept semester query from contract; tolerate missing value."""
    return (request.GET.get('semester') or 'All').strip()


def _parse_department_filter(request):
    """
    Mirrors integration doc: department=All Departments|<exact department name>.

    Loose matching is avoided to prevent substring leaks across similarly named departments.
    """
    raw = (request.GET.get('department') or '').strip()
    if not raw:
        return None
    if raw.lower() in {'all departments', 'all'}:
        return None
    return raw


def _distribution_year_bounds(year_int: int) -> bool:
    return 2000 <= year_int <= 2100


def _normalize_front_status(value: str):
    cleaned = (value or '').strip().lower()
    if cleaned in {'initial', 'pending', 'approved', 'rejected'}:
        return cleaned.upper()
    return None


def _serialize_staff_row(staff_row: Staff):
    user_obj = staff_row.user
    return {
        'id': str(staff_row.staff_id),
        'staffId': staff_row.staff_number,
        'firstName': user_obj.first_name or '',
        'lastName': user_obj.last_name or '',
        'email': user_obj.email or '',
        'title': staff_row.title or '',
        'currentDepartment': staff_row.department.name,
        'isActive': staff_row.is_active,
        'isNewEmployee': staff_row.is_new_employee,
        'notes': staff_row.notes or '',
        'updatedAt': staff_row.updated_at.strftime('%Y-%m-%d %H:%M'),
    }


def _get_distributed_time(report):
    """Return the timestamp when this report was distributed (status set to APPROVED)."""
    log = AuditLog.objects.filter(report=report, action_type='APPROVE').order_by('-created_at').first()
    return log.created_at.strftime('%Y-%m-%d %H:%M') if log else ''


def _get_operated_by(report):
    """Return the name of the staff who last approved/rejected this report."""
    log = AuditLog.objects.filter(
        report=report, action_type__in=['APPROVE', 'REJECT']
    ).select_related('action_by__user').order_by('-created_at').first()
    if not log or not log.action_by:
        return ''
    return log.action_by.user.get_full_name().strip() or log.action_by.user.username


def _serialize_workload_row(report, items):
    """Serialize a WorkloadReport to the school-operations contract list shape."""
    staff_user = report.staff.user
    full_name = staff_user.get_full_name().strip() or staff_user.username
    total_hours = sum((i.allocated_hours for i in items), Decimal('0.00'))
    first_teaching = next((i for i in items if i.category == 'TEACHING' and i.unit_code), None)
    sem = report.semester
    sem_label = f"Sem{sem[-1]}" if sem.startswith('S') else sem
    period_label = f"{report.academic_year}-{sem[-1]}" if sem.startswith('S') else str(report.academic_year)

    return {
        'id': str(report.report_id),
        'studentId': report.staff.staff_number,
        'semesterLabel': sem_label,
        'periodLabel': period_label,
        'name': full_name,
        'unit': first_teaching.unit_code if first_teaching else '',
        'notes': _get_request_reason(report),
        'requestReason': _get_request_reason(report),
        'title': report.staff.title or '',
        'department': report.snapshot_department.name,
        'rate': int(float(report.snapshot_fte) * 100),
        'status': report.status.lower(),
        'hours': _to_decimal_hours(total_hours),
        'supervisorNote': _get_supervisor_note(report),
        'operatedBy': _get_operated_by(report),
        'targetTeachingRatio': None,
        'teachingTargetHours': None,
        'cancelled': False,
        'importedFromTemplate': report.import_batch_id is not None,
        'targetBand': None,
        'workloadNewStaff': report.staff.is_new_employee,
        'hodReview': 'no',
        'distributedTime': _get_distributed_time(report),
    }


def _serialize_workload_detail(report, items):
    """Serialize a WorkloadReport to the school-operations contract detail shape."""
    staff_user = report.staff.user
    full_name = staff_user.get_full_name().strip() or staff_user.username
    total_hours = sum((i.allocated_hours for i in items), Decimal('0.00'))

    anomaly_result = evaluate_mvp_anomaly(report)
    metrics = anomaly_result['metrics']
    calc_tr = float(metrics['calc_tr'])
    calculated_band = metrics['calculated_band']

    # Build breakdown grouped by category with conflict flags
    CATEGORY_LABELS = {
        'TEACHING': 'Teaching',
        'ASSIGNED_ROLE': 'Assigned Roles',
        'HDR_SUPERVISION': 'HDR',
        'SERVICE': 'Service',
    }
    breakdown = {'Teaching': [], 'HDR': [], 'Service': [], 'Assigned Roles': [], 'Research (residual)': []}
    for item in items:
        label = CATEGORY_LABELS.get(item.category)
        if label:
            breakdown[label].append({
                'name': item.unit_code or item.description or item.category,
                'hours': _to_decimal_hours(item.allocated_hours),
            })

    research_hrs = float(metrics['research_pts']) * 17.25
    if research_hrs > 0:
        breakdown['Research (residual)'].append({
            'name': 'Research (residual)',
            'hours': round(research_hrs, 2),
        })

    failed_reasons = anomaly_result['reasons']
    return {
        'id': str(report.report_id),
        'studentId': report.staff.staff_number,
        'name': full_name,
        'department': report.snapshot_department.name,
        'status': report.status.lower(),
        'hours': _to_decimal_hours(total_hours),
        'targetTeachingRatio': None,
        'actualTeachingRatio': round(calc_tr * 100, 1),
        'targetBand': None,
        'calculatedBand': calculated_band,
        'fte': float(report.snapshot_fte),
        'workloadNewStaff': report.staff.is_new_employee,
        'hodReview': 'no',
        'cancelled': False,
        'notes': _get_request_reason(report),
        'validation': {
            'hoursAbnormal': report.is_anomaly,
            'teachingRatioWarning': 'tr_discrepancy' in failed_reasons,
            'failedReasons': failed_reasons,
        },
        'breakdown': breakdown,
    }


def _serialize_assignment_row(obj: StaffRoleAssignment):    return {
        'id': obj.assignment_id,
        'staff_id': obj.staff.staff_number,
        'role': obj.role_code,
        'department': obj.department_scope,
        'permissions': obj.permissions or [],
        'status': obj.status,
    }


def _build_visualization_payload(reports_queryset, year_from, year_to, semester_filter, dept_label_scope):
    """
    Shape aligns with frontend_api_contract_cn.md §9.11 HoS visualization for Admin reuse.

    reports_queryset must be prefetch_related('items') for performance.
    """
    departments = sorted(
        {r.snapshot_department.name for r in reports_queryset},
        key=lambda name: name.lower(),
    )

    dept_stats = {}
    for report in reports_queryset:
        dept_name = report.snapshot_department.name
        bucket = dept_stats.setdefault(dept_name, {
            'department': dept_name,
            'academics': set(),
            'total_hours': Decimal('0.00'),
            'pending': 0,
            'approved': 0,
            'rejected': 0,
        })
        bucket['academics'].add(report.staff_id)
        total_h = sum((i.allocated_hours for i in report.items.all()), Decimal('0.00'))
        bucket['total_hours'] += total_h
        status_key = report.status.lower()
        if status_key == 'pending':
            bucket['pending'] += 1
        elif status_key == 'approved':
            bucket['approved'] += 1
        elif status_key == 'rejected':
            bucket['rejected'] += 1

    department_stats = []
    for dept_name in departments:
        data = dept_stats[dept_name]
        department_stats.append({
            'department': dept_name,
            'academics': len(data['academics']),
            'total_hours': float(round(data['total_hours'], 2)),
            'pending': data['pending'],
            'approved': data['approved'],
            'rejected': data['rejected'],
        })

    trend_map = {}
    for report in reports_queryset:
        label = f"{report.academic_year} {report.semester}"
        entry = trend_map.setdefault(label, {'semester': label})
        dept_name = report.snapshot_department.name
        hrs = sum((i.allocated_hours for i in report.items.all()), Decimal('0.00'))
        entry[dept_name] = float(round(Decimal(str(entry.get(dept_name, 0))) + hrs, 2))

    workload_trend = [trend_map[key] for key in sorted(trend_map.keys())]

    total_hours_all = Decimal('0.00')
    for report in reports_queryset:
        total_hours_all += sum((i.allocated_hours for i in report.items.all()), Decimal('0.00'))
    academics_union = set()
    pending_total = approved_total = rejected_total = 0
    for report in reports_queryset:
        academics_union.add(report.staff_id)
        st = report.status.lower()
        if st == 'pending':
            pending_total += 1
        elif st == 'approved':
            approved_total += 1
        elif st == 'rejected':
            rejected_total += 1

    reporting_period_label = f"{year_from or '?'}-{year_to or '?'}"
    semester_part = semester_filter.upper() if semester_filter else 'ALL'
    if semester_part == 'ALL':
        reporting_period_label += ' All Semesters'
    else:
        reporting_period_label += f' {semester_part}'
    scope_label = dept_label_scope or 'All Departments'

    return {
        'reportingPeriodLabel': reporting_period_label,
        'scopeLabel': scope_label,
        'summary': {
            'totalDepartments': len(departments),
            'totalAcademics': len(academics_union),
            'totalWorkHours': float(round(total_hours_all, 2)),
            'pendingRequests': pending_total,
            'approvedRequests': approved_total,
            'rejectedRequests': rejected_total,
        },
        'departmentStats': department_stats,
        'trend': workload_trend,
    }


def _staff_from_body_or_path(request, lookup_id: str):
    """Resolve staff rows using immutable staff_number identifiers from the contracts."""
    return get_object_or_404(Staff, staff_number=lookup_id.strip())


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_workload_requests(request):
    """GET /api/school-operations/workloads  (also /api/admin/workload-requests/)"""
    base_qs = _admin_reports_qs(request.staff).prefetch_related('items').select_related(
        'staff__user', 'staff__department', 'snapshot_department'
    )

    qs = base_qs

    # New contract uses status_filter; legacy used status — accept both.
    status_filter = (request.GET.get('status_filter') or request.GET.get('status') or 'all').lower()
    if status_filter == 'pending':
        qs = qs.filter(status='PENDING')
    elif status_filter == 'distributed':
        qs = qs.filter(status='APPROVED')
    elif status_filter == 'failed':
        qs = qs.filter(status='REJECTED')
    elif status_filter == 'superseded':
        # superseded = non-current; override base_qs which already filters is_current=True
        qs = get_workload_queryset(request.staff).filter(is_current=False).prefetch_related('items').select_related(
            'staff__user', 'staff__department', 'snapshot_department'
        )
    elif status_filter == 'initial':
        qs = qs.filter(status='INITIAL')
    # 'all' → no additional filter

    # New contract query params
    staff_id = request.GET.get('staff_id', '').strip()
    if staff_id:
        qs = qs.filter(staff__staff_number=staff_id)

    name = request.GET.get('name', '').strip()
    if name:
        qs = qs.filter(
            models.Q(staff__user__first_name__icontains=name)
            | models.Q(staff__user__last_name__icontains=name)
        )

    dept_name = _parse_department_filter(request)
    if dept_name:
        qs = qs.filter(snapshot_department__name=dept_name)

    year = request.GET.get('year', '').strip()
    if year:
        qs = qs.filter(academic_year=year)

    semester = request.GET.get('semester', '').strip()
    if semester and semester.upper() != 'ALL':
        qs = qs.filter(semester=semester.upper())

    qs = qs.order_by('-updated_at')

    counts = {
        'pending': base_qs.filter(status='PENDING').count(),
        'distributed': base_qs.filter(status='APPROVED').count(),
        'failed': base_qs.filter(status='REJECTED').count(),
        'superseded': get_workload_queryset(request.staff).filter(is_current=False).count(),
    }

    try:
        page = max(1, int(request.GET.get('page', 1)))
        page_size = max(1, min(100, int(request.GET.get('page_size', 10))))
    except (ValueError, TypeError):
        return Response(
            {'success': False, 'message': 'page and page_size must be positive integers'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    rows = [_serialize_workload_row(r, list(r.items.all())) for r in qs]
    paginator = Paginator(rows, page_size)
    current_page = paginator.get_page(page)

    return Response({
        'success': True,
        'message': 'Workload list loaded',
        'data': {
            'items': list(current_page.object_list),
            'pagination': {
                'page': current_page.number,
                'pageSize': page_size,
                'totalItems': paginator.count,
                'totalPages': paginator.num_pages,
            },
            'counts': counts,
        },
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_workload_request_detail(request, id):
    """GET /api/school-operations/workloads/{id}  (also /api/admin/workload-requests/{id}/)"""
    qs = (
        _admin_reports_qs(request.staff)
        .prefetch_related('items')
        .select_related('staff__user', 'staff__department', 'snapshot_department')
    )
    report = get_object_or_404(qs, report_id=id)
    items = list(report.items.all())

    return Response({
        'success': True,
        'message': 'Workload detail loaded',
        'data': _serialize_workload_detail(report, items),
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@transaction.atomic
def admin_batch_decision(request):
    """POST /api/admin/workload-requests/batch-decision/"""
    request_ids = request.data.get('request_ids') or []
    decision = (request.data.get('decision') or '').lower()

    if not isinstance(request_ids, list) or not request_ids:
        return Response(
            {'success': False, 'message': 'Validation failed',
             'errors': {'request_ids': ['At least one id is required']}},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    if len(request_ids) > 100:
        return Response(
            {'success': False, 'message': 'Validation failed',
             'errors': {'request_ids': ['Cannot process more than 100 ids at once']}},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    if decision not in ('approved', 'rejected'):
        return Response(
            {'success': False, 'message': 'Validation failed',
             'errors': {'decision': ['Must be approved or rejected']}},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    qs = _admin_reports_qs(request.staff)
    reports = list(qs.filter(report_id__in=request_ids))

    if len(reports) != len(request_ids):
        found_ids = {str(r.report_id) for r in reports}
        missing_ids = [rid for rid in request_ids if str(rid) not in found_ids]
        stale_ids = list(
            WorkloadReport.objects.filter(report_id__in=missing_ids, is_current=False)
            .values_list('report_id', flat=True)
        )
        if stale_ids:
            return Response(
                {**STALE_REPORT_ERROR, 'errors': {'request_ids': [str(x) for x in stale_ids]}},
                status=http_status.HTTP_409_CONFLICT,
            )
        return Response(
            {'success': False, 'message': 'One or more request ids are invalid or not accessible'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    non_pending = [r for r in reports if r.status != 'PENDING']
    if non_pending:
        return Response(
            {'success': False, 'message': 'One or more requests are not in PENDING status',
             'errors': {'request_ids': [str(r.report_id) for r in non_pending]}},
            status=http_status.HTTP_409_CONFLICT,
        )

    action_type = 'APPROVE' if decision == 'approved' else 'REJECT'
    new_status = decision.upper()
    for report in reports:
        report.status = new_status
        report.save(update_fields=['status', 'updated_at'])
        AuditLog.objects.create(report=report, action_by=request.staff, action_type=action_type)

    return Response({
        'success': True,
        'message': 'Batch decision completed',
        'data': {'updated_count': len(reports), 'decision': decision},
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@transaction.atomic
def admin_single_decision(request, id):
    """POST /api/admin/workload-requests/{id}/decision/"""
    decision = (request.data.get('decision') or '').lower()
    note = (request.data.get('note') or '').strip()
    breakdown_data = request.data.get('breakdown')

    if decision not in ('approved', 'rejected'):
        return Response(
            {'success': False, 'message': 'Validation failed',
             'errors': {'decision': ['Must be approved or rejected']}},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    if not note:
        return Response(
            {'success': False, 'message': 'Validation failed',
             'errors': {'note': ['note is required']}},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    if len(note) > 240:
        return Response(
            {'success': False, 'message': 'Validation failed',
             'errors': {'note': ['note must be <= 240 characters']}},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    qs = _admin_reports_qs(request.staff)
    report = qs.filter(report_id=id).first()
    if report is None:
        stale = WorkloadReport.objects.filter(report_id=id, is_current=False).first()
        if stale is not None:
            return Response(STALE_REPORT_ERROR, status=http_status.HTTP_409_CONFLICT)
        return Response(
            {'success': False, 'message': 'Report not found'},
            status=http_status.HTTP_404_NOT_FOUND,
        )

    if report.status != 'PENDING':
        return Response(
            {'success': False, 'message': f"Report status is '{report.status}', must be PENDING"},
            status=http_status.HTTP_409_CONFLICT,
        )

    if breakdown_data and isinstance(breakdown_data, dict):
        parsed, parse_errors = _parse_breakdown_data(breakdown_data)
        if parse_errors:
            return Response(
                {'success': False, 'message': 'Validation failed', 'errors': {'breakdown': parse_errors}},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        if not parsed:
            return Response(
                {'success': False, 'message': 'Validation failed',
                 'errors': {'breakdown': ['At least one valid breakdown row is required']}},
                status=http_status.HTTP_400_BAD_REQUEST,
            )
        report.items.all().delete()
        WorkloadItem.objects.bulk_create([
            WorkloadItem(report=report, **kwargs) for kwargs in parsed
        ])

    action_type = 'APPROVE' if decision == 'approved' else 'REJECT'
    report.status = decision.upper()
    report.save(update_fields=['status', 'updated_at'])
    AuditLog.objects.create(report=report, action_by=request.staff, action_type=action_type, comment=note)

    return Response({
        'success': True,
        'message': 'Request updated',
        'data': {
            'id': str(report.report_id),
            'status': decision,
            'supervisor_note': note,
        },
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@transaction.atomic
def admin_distribute_workloads(request):
    """POST /api/school-operations/workloads/distribute  (also /api/admin/workloads/distribute/)"""
    # New contract: workloadIds + academicYear + semester
    workload_ids = request.data.get('workloadIds') or []
    year = request.data.get('academicYear') or request.data.get('year')
    semester = (request.data.get('semester') or '').strip().upper()

    try:
        year_int = int(year)
    except (TypeError, ValueError):
        return Response(
            {'success': False, 'message': 'academicYear must be a valid integer'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    if not _distribution_year_bounds(year_int):
        return Response(
            {'success': False, 'message': 'academicYear outside allowed range (2000-2100)'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    if semester not in {'S1', 'S2'}:
        return Response(
            {'success': False, 'message': 'semester must be S1 or S2'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    if not isinstance(workload_ids, list) or not workload_ids:
        return Response(
            {'success': False, 'message': 'workloadIds must be a non-empty list'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    if len(workload_ids) > 200:
        return Response(
            {'success': False, 'message': 'Cannot distribute more than 200 workloads at once'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    qs = _admin_reports_qs(request.staff).select_related('staff__user')
    reports = list(qs.filter(report_id__in=workload_ids))

    if len(reports) != len(workload_ids):
        return Response(
            {'success': False, 'message': 'One or more workloadIds are invalid or not accessible'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    job = WorkloadDistributionJob.objects.create(
        academic_year=year_int,
        semester=semester,
        triggered_by=request.staff,
        notes=f'Distributed {len(reports)} workloads via school-operations portal',
    )

    now_str = timezone.now().strftime('%Y-%m-%d %H:%M')
    operated_by = request.staff.user.get_full_name().strip() or request.staff.user.username
    items_out = []
    for report in reports:
        items_out.append({
            'workloadId': str(report.report_id),
            'status': report.status.lower(),
            'distributedTime': now_str,
            'operatedBy': operated_by,
        })

    return Response({
        'success': True,
        'message': 'Workload distribution job created',
        'data': {
            'processedCount': len(reports),
            'jobId': job.job_id,
            'items': items_out,
        },
    }, status=http_status.HTTP_201_CREATED)


def _write_workbook(headers, rows):
    try:
        import openpyxl
    except ImportError:
        return None

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _materialize_template(relative_path: Path, headers, sample_rows=None):
    buffer = _write_workbook(headers, sample_rows or [])
    if buffer is None:
        return False, None

    destination = relative_path
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(buffer.getvalue())
    return True, destination


def _admin_template_urls(request, subpath_suffix: str, filename: str, headers, sample_rows):
    """
    Returns JSON contract plus real download endpoints for environments without public MEDIA access.
    """
    media_dir = _ensure_media_subdir(TEMPLATE_MEDIA_SUBDIR)
    target = media_dir / filename
    ok, disk_path = _materialize_template(target, headers, sample_rows)
    if not ok:
        return Response(
            {'success': False, 'message': 'Template unavailable: openpyxl not installed'},
            status=http_status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    download_url = request.build_absolute_uri(f'/api/admin/{subpath_suffix}/download/')
    return Response({
        'success': True,
        'message': 'Template ready',
        'data': {
            'file_name': filename,
            'download_url': download_url,
        },
    })


def _dispatch_template_download(request, filename: str):
    media_dir = _ensure_media_subdir(TEMPLATE_MEDIA_SUBDIR)
    target = media_dir / filename
    if not target.exists():
        return Response({'success': False, 'message': 'Template missing; regenerate listing first.'}, status=404)
    safe_name = re.sub(r'[^A-Za-z0-9_.-]', '_', filename)
    # Read into memory first to avoid Windows file-handle lock (WinError 32).
    payload = target.read_bytes()
    response = HttpResponse(payload, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{safe_name}"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_workload_import_template(request):
    headers = ['employee_id', 'name', 'description', 'total_work_hours', 'status']
    return _admin_template_urls(request, 'workloads/import-template', 'Workload_Template.xlsx', headers, [])


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_workload_import_template_download(request):
    return _dispatch_template_download(request, 'Workload_Template.xlsx')


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_staff_import_template(request):
    headers = ['employee_id', 'first_name', 'last_name', 'email', 'department', 'active_status']
    return _admin_template_urls(request, 'staff/import-template', 'Staff_Template.xlsx', headers, [])


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_staff_import_template_download(request):
    return _dispatch_template_download(request, 'Staff_Template.xlsx')


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@throttle_classes([AdminImportThrottle])
@transaction.atomic
def admin_workload_import(request):
    """POST /api/school-operations/workloads/import  (also /api/admin/workloads/import/)

    Accepts JSON body from the frontend (browser-parsed workbook data).
    The old Excel file-upload path is no longer the primary interface.
    """
    body = request.data or {}
    sheets = body.get('sheets')
    if not isinstance(sheets, list) or not sheets:
        return Response(
            {'success': False, 'message': 'sheets must be a non-empty list'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    batch_id = uuid.uuid4()
    created_count = 0
    updated_count = 0
    failed_count = 0
    failures = []

    for sheet in sheets:
        sheet_name = sheet.get('sheetName', '')
        # Infer semester from sheet name: "Sem1" → S1, "Sem2" → S2
        sem_raw = str(sheet_name).strip()
        if '1' in sem_raw:
            semester = 'S1'
        elif '2' in sem_raw:
            semester = 'S2'
        else:
            semester = 'S1'

        anomaly_metrics = sheet.get('anomalyMetricsByStaffId') or {}
        teaching_lines = sheet.get('teachingLinesByStaffId') or {}
        hdr_metrics = sheet.get('hdrMetricsByStaffId') or {}
        service_metrics = sheet.get('serviceMetricsByStaffId') or {}
        role_metrics = sheet.get('roleMetricsByStaffId') or {}

        # Collect all staff IDs from this sheet
        all_staff_ids = set(anomaly_metrics.keys()) | set(teaching_lines.keys())

        for staff_number in all_staff_ids:
            staff_row = Staff.objects.select_related('department', 'user').filter(
                staff_number=staff_number
            ).first()
            if not staff_row:
                failures.append({'staffId': staff_number, 'sheet': sheet_name, 'message': 'Staff not found'})
                failed_count += 1
                continue

            metrics = anomaly_metrics.get(staff_number) or {}
            year_val = body.get('importedAtIso', '')[:4]
            try:
                year_int = int(year_val)
            except (ValueError, TypeError):
                year_int = timezone.now().year

            # Block re-import if a non-INITIAL/REJECTED report already exists
            conflicts = WorkloadReport.objects.filter(
                staff=staff_row,
                academic_year=year_int,
                semester=semester,
                is_current=True,
            ).exclude(status__in=['INITIAL', 'REJECTED'])
            if conflicts.exists():
                failures.append({'staffId': staff_number, 'sheet': sheet_name, 'message': 'Report locked; rollback required'})
                failed_count += 1
                continue

            orphan_reports = list(WorkloadReport.objects.select_for_update().filter(
                staff=staff_row,
                academic_year=year_int,
                semester=semester,
                is_current=True,
            ))

            # Always force INITIAL — import must never bypass the approval workflow.
            report = WorkloadReport.objects.create(
                staff=staff_row,
                academic_year=year_int,
                semester=semester,
                snapshot_fte=staff_row.fte,
                snapshot_department=staff_row.department,
                status='INITIAL',
                import_batch_id=batch_id,
                is_anomaly=False,
                is_current=True,
            )

            for old in orphan_reports:
                old.is_current = False
                old.superseded_by = report
                old.save(update_fields=['is_current', 'superseded_by', 'updated_at'])
                AuditLog.objects.create(
                    report=old,
                    action_by=request.staff,
                    action_type='MODIFIED_BY_REIMPORT',
                    changes={'superseded_by': str(report.report_id), 'batch': str(batch_id)},
                )

            AuditLog.objects.create(
                report=report,
                action_by=request.staff,
                action_type='IMPORTED',
                changes={'batch': str(batch_id), 'kind': 'JSON_WORKLOAD_IMPORT',
                         'superseded': [str(r.report_id) for r in orphan_reports]},
            )

            # Create WorkloadItems from the parsed sheet data
            items_to_create = []

            for line in (teaching_lines.get(staff_number) or []):
                hrs = Decimal(str(line.get('hours', 0) or 0))
                if hrs < 0:
                    continue
                items_to_create.append(WorkloadItem(
                    report=report,
                    category='TEACHING',
                    unit_code=str(line.get('unit', ''))[:50] or None,
                    description='Teaching',
                    allocated_hours=hrs,
                ))

            hdr = hdr_metrics.get(staff_number) or {}
            hdr_hrs = Decimal(str(hdr.get('totalHrs', 0) or 0))
            if hdr_hrs > 0:
                items_to_create.append(WorkloadItem(
                    report=report,
                    category='HDR_SUPERVISION',
                    unit_code=None,
                    description='HDR Supervision',
                    allocated_hours=hdr_hrs,
                ))

            svc = service_metrics.get(staff_number) or {}
            svc_pts = Decimal(str(svc.get('servicePoints', 0) or 0))
            svc_hrs = svc_pts * Decimal('17.25')
            if svc_hrs > 0:
                items_to_create.append(WorkloadItem(
                    report=report,
                    category='SERVICE',
                    unit_code=None,
                    description='Service',
                    allocated_hours=svc_hrs,
                ))

            for role in (role_metrics.get(staff_number) or {}).get('roles', []):
                role_hrs = Decimal(str(role.get('hours', 0) or 0))
                if role_hrs > 0:
                    items_to_create.append(WorkloadItem(
                        report=report,
                        category='ASSIGNED_ROLE',
                        unit_code=None,
                        description=str(role.get('name', 'Role'))[:500],
                        allocated_hours=role_hrs,
                    ))

            if items_to_create:
                WorkloadItem.objects.bulk_create(items_to_create)

            persist_report_anomaly(report, department_conflict=False)

            if orphan_reports:
                updated_count += 1
            else:
                created_count += 1

    return Response({
        'ok': True,
        'referenceId': str(batch_id),
        'created': created_count,
        'updated': updated_count,
        'failed': failed_count,
        'errors': failures,
    })


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@throttle_classes([AdminImportThrottle])
@transaction.atomic
def admin_staff_import(request):
    """POST /api/school-operations/staff/import  (also /api/admin/staff/import/)

    Accepts JSON body: { "rows": [ { staffId, firstName, lastName, email, title,
    department, isActive, isNewEmployee, notes } ] }
    Updates existing Staff only — never creates phantom users.
    """
    body = request.data or {}
    rows = body.get('rows')
    if not isinstance(rows, list) or not rows:
        return Response(
            {'success': False, 'message': 'rows must be a non-empty list'},
            status=http_status.HTTP_400_BAD_REQUEST,
        )

    created_count = 0
    updated_count = 0
    failures = []

    for idx, row in enumerate(rows):
        staff_number = str(row.get('staffId', '')).strip()
        if not staff_number:
            failures.append({'index': idx, 'message': 'staffId is required'})
            continue

        staff_row = Staff.objects.select_related('user', 'department').filter(staff_number=staff_number).first()
        if not staff_row:
            failures.append({'index': idx, 'staffId': staff_number, 'message': 'Staff not found'})
            continue

        user_obj = staff_row.user
        user_fields = []

        first_name = row.get('firstName')
        if first_name is not None:
            user_obj.first_name = str(first_name).strip()[:150]
            user_fields.append('first_name')

        last_name = row.get('lastName')
        if last_name is not None:
            user_obj.last_name = str(last_name).strip()[:150]
            user_fields.append('last_name')

        email = row.get('email')
        if email is not None:
            email_clean = str(email).strip().lower()
            if email_clean and not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email_clean):
                failures.append({'index': idx, 'staffId': staff_number, 'message': 'invalid email'})
                continue
            user_obj.email = email_clean
            user_fields.append('email')

        if user_fields:
            user_obj.save(update_fields=list(set(user_fields)))

        staff_fields = []

        dept_name = row.get('department')
        if dept_name:
            dept = Department.objects.filter(name__iexact=str(dept_name).strip()).first()
            if not dept:
                failures.append({'index': idx, 'staffId': staff_number, 'message': 'department not found'})
                continue
            staff_row.department = dept
            staff_fields.append('department')

        title = row.get('title')
        if title is not None:
            staff_row.title = str(title).strip()[:100]
            staff_fields.append('title')

        is_active = row.get('isActive')
        if is_active is not None:
            staff_row.is_active = bool(is_active)
            staff_fields.append('is_active')

        is_new = row.get('isNewEmployee')
        if is_new is not None:
            staff_row.is_new_employee = bool(is_new)
            staff_fields.append('is_new_employee')

        notes = row.get('notes')
        if notes is not None:
            staff_row.notes = str(notes)
            staff_fields.append('notes')

        if staff_fields:
            staff_fields.append('updated_at')
            staff_row.save(update_fields=staff_fields)

        updated_count += 1

    return Response({
        'ok': True,
        'created': created_count,
        'updated': updated_count,
        'errors': failures,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_staff_list(request):
    """GET /api/school-operations/staff  (also /api/admin/staff/)"""
    queryset = Staff.objects.select_related('user', 'department').order_by('staff_number')

    # New contract query params
    staff_id = request.GET.get('staff_id', '').strip()
    if staff_id:
        queryset = queryset.filter(staff_number=staff_id)

    first_name = request.GET.get('first_name', '').strip()
    if first_name:
        queryset = queryset.filter(user__first_name__icontains=first_name)

    last_name = request.GET.get('last_name', '').strip()
    if last_name:
        queryset = queryset.filter(user__last_name__icontains=last_name)

    # Legacy search param
    search_term = request.GET.get('query', '').strip()
    if search_term:
        queryset = queryset.filter(
            models.Q(user__first_name__icontains=search_term)
            | models.Q(user__last_name__icontains=search_term)
            | models.Q(staff_number__icontains=search_term)
        )

    try:
        page = max(1, int(request.GET.get('page', 1)))
        page_size = max(1, min(100, int(request.GET.get('page_size', 10))))
    except (ValueError, TypeError):
        return Response({'success': False, 'message': 'invalid pagination'}, status=400)

    paginator = Paginator(queryset, page_size)
    page_obj = paginator.get_page(page)

    return Response({
        'success': True,
        'message': 'Staff roster loaded',
        'data': {
            'items': [_serialize_staff_row(s) for s in page_obj.object_list],
            'pagination': {
                'page': page_obj.number,
                'pageSize': page_size,
                'totalItems': paginator.count,
                'totalPages': paginator.num_pages,
            },
        },
    })


@api_view(['GET', 'PATCH'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@transaction.atomic
def admin_staff_patch(request, staff_id):
    """GET /api/school-operations/staff/{staffId}  or  PATCH /api/school-operations/staff/{staffId}"""
    staff_row = _staff_from_body_or_path(request, staff_id)

    if request.method == 'GET':
        return Response({
            'success': True,
            'message': 'Staff detail loaded',
            'data': _serialize_staff_row(staff_row),
        })

    # PATCH — accept both camelCase (new contract) and snake_case (legacy)
    payload = request.data or {}

    first_name = payload.get('firstName') if 'firstName' in payload else payload.get('first_name')
    last_name = payload.get('lastName') if 'lastName' in payload else payload.get('last_name')
    email = payload.get('email')
    dept_name = payload.get('department')
    title = payload.get('title')
    is_active = payload.get('isActive') if 'isActive' in payload else (
        None if 'active_status' not in payload else (payload.get('active_status', '').lower() != 'inactive')
    )
    is_new_employee = payload.get('isNewEmployee')
    notes = payload.get('notes')

    user_obj = staff_row.user
    user_fields = []

    if first_name is not None:
        user_obj.first_name = str(first_name).strip()[:150]
        user_fields.append('first_name')
    if last_name is not None:
        user_obj.last_name = str(last_name).strip()[:150]
        user_fields.append('last_name')
    if email is not None:
        email_clean = str(email).strip().lower()
        if not re.match(r'^[^@\s]+@[^@\s]+\.[^@\s]+$', email_clean):
            return Response({'success': False, 'message': 'invalid email'}, status=http_status.HTTP_400_BAD_REQUEST)
        user_obj.email = email_clean
        user_fields.append('email')

    if user_fields:
        user_obj.save(update_fields=list(set(user_fields)))

    staff_fields = []
    if dept_name:
        department = Department.objects.filter(name__iexact=str(dept_name).strip()).first()
        if not department:
            return Response({'success': False, 'message': 'department not found'}, status=http_status.HTTP_400_BAD_REQUEST)
        staff_row.department = department
        staff_fields.append('department')

    if title is not None:
        staff_row.title = str(title).strip()[:100]
        staff_fields.append('title')

    if is_active is not None:
        staff_row.is_active = bool(is_active)
        staff_fields.append('is_active')

    if is_new_employee is not None:
        staff_row.is_new_employee = bool(is_new_employee)
        staff_fields.append('is_new_employee')

    if notes is not None:
        staff_row.notes = str(notes)
        staff_fields.append('notes')

    if staff_fields:
        staff_fields.append('updated_at')
        staff_row.save(update_fields=staff_fields)

    return Response({
        'ok': True,
        'staff': _serialize_staff_row(staff_row),
    })


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_role_assignments(request):
    """GET list + POST create."""
    if request.method == 'GET':
        queryset = StaffRoleAssignment.objects.select_related('staff', 'resolved_department').order_by('-created_at')
        return Response({'success': True, 'message': 'Assignments loaded', 'data': {
            'items': [_serialize_assignment_row(obj) for obj in queryset[:500]],
        }})

    body = request.data or {}
    staff_number = str(body.get('staff_id', '')).strip()
    role_front = body.get('role')
    dept_scope = str(body.get('department', '')).strip()
    permissions = body.get('permissions') or []

    if not staff_number or role_front not in dict(StaffRoleAssignment.FRONT_ROLE_CHOICES):
        return Response({'success': False, 'message': 'staff_id and role are required'}, status=http_status.HTTP_400_BAD_REQUEST)
    if not isinstance(permissions, list):
        return Response({'success': False, 'message': 'permissions must be a list'}, status=http_status.HTTP_400_BAD_REQUEST)

    staff_row = Staff.objects.filter(staff_number=staff_number).first()
    if not staff_row:
        return Response({'success': False, 'message': 'staff not found'}, status=http_status.HTTP_404_NOT_FOUND)

    resolved = Department.objects.filter(name__iexact=dept_scope).first()

    assignment = StaffRoleAssignment.objects.create(
        staff=staff_row,
        role_code=role_front,
        department_scope=dept_scope,
        resolved_department=resolved,
        permissions=permissions,
        status='active',
    )

    # Sync Staff.role so require_role() checks take effect immediately.
    _FRONT_TO_CANONICAL = {'HoD': 'HOD', 'Admin': 'SCHOOL_OPS'}
    canonical = _FRONT_TO_CANONICAL.get(role_front)
    if canonical:
        staff_row.role = canonical
        staff_row.save(update_fields=['role', 'updated_at'])

    return Response({
        'success': True,
        'message': 'Role assigned',
        'data': {
            **_serialize_assignment_row(assignment),
        },
    }, status=http_status.HTTP_201_CREATED)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@transaction.atomic
def admin_role_assignment_disable(request, assignment_id):
    payload = request.data or {}
    reason = str(payload.get('reason') or '').strip()

    assignment = get_object_or_404(StaffRoleAssignment, assignment_id=int(assignment_id))
    assignment.status = 'disabled'
    assignment.disable_reason = reason[:500]
    assignment.save(update_fields=['status', 'disable_reason', 'updated_at'])

    return Response({'success': True, 'message': 'Role assignment disabled', 'data': {
        'id': assignment.assignment_id,
        'status': assignment.status,
    }})


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_visualization(request):
    semester_filter = _parse_semester_filter(request)
    year_from, year_to = _parse_year_range(request)
    dept_scope = _parse_department_filter(request)

    queryset = (
        _admin_reports_qs(request.staff)
        .prefetch_related('items')
        .select_related('snapshot_department', 'staff__user')
    )
    queryset = _filter_reports_by_range(queryset, year_from, year_to, semester_filter)
    if dept_scope:
        queryset = queryset.filter(snapshot_department__name=dept_scope)

    payload = _build_visualization_payload(list(queryset), year_from, year_to, semester_filter, dept_scope)

    return Response({'success': True, 'message': 'Visualization loaded', 'data': payload})


def _persist_export_workbook(request):
    """Write Excel to disk and register cache token for owner-only download."""
    try:
        import openpyxl
    except ImportError:
        return None, Response(
            {'success': False, 'message': 'Export unavailable: openpyxl not installed'},
            status=http_status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    year_from, year_to = _parse_year_range(request)
    semester_filter = _parse_semester_filter(request)
    dept_scope = _parse_department_filter(request)

    queryset = (
        _admin_reports_qs(request.staff)
        .prefetch_related('items')
        .select_related('staff__user', 'snapshot_department')
        .filter(is_current=True)
        .order_by('snapshot_department__name', 'staff__staff_number', 'academic_year', 'semester')
    )
    queryset = _filter_reports_by_range(queryset, year_from, year_to, semester_filter)
    if dept_scope:
        queryset = queryset.filter(snapshot_department__name=dept_scope)

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = 'Admin Export'
    headers = ['Staff Number', 'Name', 'Department', 'Year', 'Semester', 'Status', 'Hours', 'Category', 'Detail']
    sheet.append(headers)

    export_stamp = timezone.now().isoformat(timespec='seconds')
    total_rows = 0
    for report in queryset:
        staff_user = report.staff.user
        name = staff_user.get_full_name().strip() or staff_user.username
        dept = report.snapshot_department.name
        hours_total = sum((i.allocated_hours for i in report.items.all()), Decimal('0.00'))

        sheet.append([
            report.staff.staff_number,
            name,
            dept,
            report.academic_year,
            report.semester,
            report.status.lower(),
            float(hours_total),
            '',
            '',
        ])
        for item in report.items.all():
            sheet.append([
                report.staff.staff_number,
                name,
                dept,
                report.academic_year,
                report.semester,
                report.status.lower(),
                float(item.allocated_hours),
                item.category,
                item.unit_code or item.description or '',
            ])
            total_rows += 1

    export_dir = _ensure_media_subdir(EXPORT_MEDIA_SUBDIR)
    token = uuid.uuid4().hex
    fname = Path(f'{token}_Admin_Workload.xlsx')
    buffer = io.BytesIO()
    workbook.save(buffer)
    disk_path = export_dir / fname
    disk_path.write_bytes(buffer.getvalue())

    meta = {'relative': str(fname), 'issued_at': export_stamp}
    cache.set(f'admin_export:{token}', {'staff_uuid': str(request.staff.pk), 'payload': meta}, timeout=900)
    download_url = request.build_absolute_uri(f'/api/admin/export/download/?token={token}')
    return {
        'file_name': 'Admin_Workload.xlsx',
        'download_url': download_url,
        'issued_at': export_stamp,
        'rows_written': total_rows,
    }, None


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@throttle_classes([AdminExportThrottle])
def admin_export_manifest(request):
    """
    Primary contract endpoint returning JSON pointers.

    If cai adjusts export semantics later, swap the implementation behind `_persist_export_workbook`.
    """
    snapshot, error_response = _persist_export_workbook(request)
    if error_response:
        return error_response

    return Response({
        'success': True,
        'message': 'Export prepared',
        'data': snapshot,
    })


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@throttle_classes([AdminExportThrottle])
def admin_export_download(request):
    """Binary companion for `/admin/export/` JSON contracts."""
    token = request.GET.get('token')
    entry = cache.get(f'admin_export:{token}')
    if not token or not entry:
        return Response({'success': False, 'message': 'Invalid or expired token'}, status=http_status.HTTP_404_NOT_FOUND)
    if str(entry['staff_uuid']) != str(request.staff.pk):
        return Response({'success': False, 'message': 'Token does not belong to this user'}, status=http_status.HTTP_403_FORBIDDEN)

    fname = Path(entry['payload']['relative'])
    export_dir = _ensure_media_subdir(EXPORT_MEDIA_SUBDIR)
    disk_path = export_dir / fname.name

    if not disk_path.exists():
        return Response({'success': False, 'message': 'File missing'}, status=http_status.HTTP_410_GONE)

    payload_bytes = disk_path.read_bytes()
    disk_path.unlink(missing_ok=True)
    cache.delete(f'admin_export:{token}')

    response = HttpResponse(
        payload_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="Admin_Workload.xlsx"'
    return response


def _build_workload_export_workbook(qs):
    """Build an in-memory Excel workbook from a WorkloadReport queryset."""
    try:
        import openpyxl
    except ImportError:
        return None
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Workloads'
    ws.append(['Staff ID', 'Name', 'Department', 'Year', 'Semester', 'Status', 'Hours', 'Category', 'Detail'])
    for report in qs:
        staff_user = report.staff.user
        name = staff_user.get_full_name().strip() or staff_user.username
        for item in report.items.all():
            ws.append([
                report.staff.staff_number,
                name,
                report.snapshot_department.name,
                report.academic_year,
                report.semester,
                report.status.lower(),
                float(item.allocated_hours),
                item.category,
                item.unit_code or item.description or '',
            ])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@throttle_classes([AdminExportThrottle])
def admin_workload_export(request):
    """GET /api/school-operations/workloads/export — direct file stream filtered by status/staff/dept/year/semester."""
    status_filter = (request.GET.get('status_filter') or '').lower()
    qs = (
        _admin_reports_qs(request.staff)
        .prefetch_related('items')
        .select_related('staff__user', 'snapshot_department')
    )
    if status_filter == 'distributed':
        qs = qs.filter(status='APPROVED')
    elif status_filter == 'failed':
        qs = qs.filter(status='REJECTED')
    elif status_filter == 'superseded':
        qs = get_workload_queryset(request.staff).filter(is_current=False).prefetch_related('items').select_related(
            'staff__user', 'snapshot_department'
        )

    staff_id = request.GET.get('staff_id', '').strip()
    if staff_id:
        qs = qs.filter(staff__staff_number=staff_id)

    name = request.GET.get('name', '').strip()
    if name:
        qs = qs.filter(
            models.Q(staff__user__first_name__icontains=name)
            | models.Q(staff__user__last_name__icontains=name)
        )

    dept_name = _parse_department_filter(request)
    if dept_name:
        qs = qs.filter(snapshot_department__name=dept_name)

    year = request.GET.get('year', '').strip()
    if year:
        qs = qs.filter(academic_year=year)

    semester = request.GET.get('semester', '').strip()
    if semester and semester.upper() != 'ALL':
        qs = qs.filter(semester=semester.upper())

    payload = _build_workload_export_workbook(qs)
    if payload is None:
        return Response({'success': False, 'message': 'Export unavailable: openpyxl not installed'}, status=503)

    response = HttpResponse(payload, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Workloads_Export.xlsx"'
    return response


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
@throttle_classes([AdminExportThrottle])
def admin_school_export(request):
    """GET /api/school-operations/export — school-level history Excel, direct file stream."""
    year_from, year_to = _parse_year_range(request)
    semester_filter = _parse_semester_filter(request)
    dept_scope = _parse_department_filter(request)

    qs = (
        _admin_reports_qs(request.staff)
        .prefetch_related('items')
        .select_related('staff__user', 'snapshot_department')
        .order_by('snapshot_department__name', 'staff__staff_number', 'academic_year', 'semester')
    )
    qs = _filter_reports_by_range(qs, year_from, year_to, semester_filter)
    if dept_scope:
        qs = qs.filter(snapshot_department__name=dept_scope)

    payload = _build_workload_export_workbook(qs)
    if payload is None:
        return Response({'success': False, 'message': 'Export unavailable: openpyxl not installed'}, status=503)

    response = HttpResponse(payload, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="School_Workload_History.xlsx"'
    return response


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role(*ADMIN_ROLES)
def admin_contact_staff(request):
    """POST /api/school-operations/contact-staff — stub; stores message as AuditLog comment."""
    body = request.data or {}
    recipient_id = str(body.get('recipientStaffId', '')).strip()
    message_body = str(body.get('messageBody', '')).strip()

    if not recipient_id:
        return Response({'success': False, 'message': 'recipientStaffId is required'}, status=http_status.HTTP_400_BAD_REQUEST)
    if not message_body:
        return Response({'success': False, 'message': 'messageBody is required'}, status=http_status.HTTP_400_BAD_REQUEST)
    if len(message_body) > 2000:
        return Response({'success': False, 'message': 'messageBody must be <= 2000 characters'}, status=http_status.HTTP_400_BAD_REQUEST)

    recipient = Staff.objects.filter(staff_number=recipient_id).first()
    if not recipient:
        return Response({'success': False, 'message': 'Recipient staff not found'}, status=http_status.HTTP_404_NOT_FOUND)

    ref_id = f'msg_{uuid.uuid4().hex[:8]}'
    # Store as an audit entry on the recipient's most recent current report (best-effort).
    latest_report = WorkloadReport.objects.filter(staff=recipient, is_current=True).order_by('-updated_at').first()
    if latest_report:
        AuditLog.objects.create(
            report=latest_report,
            action_by=request.staff,
            action_type='CONTACT_STAFF',
            comment=message_body,
            changes={'referenceId': ref_id, 'recipientStaffId': recipient_id},
        )

    return Response({'ok': True, 'referenceId': ref_id})
