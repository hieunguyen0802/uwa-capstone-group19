import csv
import io
from datetime import date
from decimal import Decimal

from django.contrib.auth.models import User
from django.core.paginator import Paginator
from django.db import transaction
from django.db.models import Exists, OuterRef, Q
from django.http import HttpResponse
from django.shortcuts import get_object_or_404
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from api.decorators import require_role
from api.models import AuditLog, Department, Staff, WorkloadReport
from api.services.workload_service import (
    _filter_reports_by_range,
    _parse_year_range,
    _reporting_period_label,
)

ROLE_ALIAS_TO_DB = {
    'HoD': 'HOD',
    'Admin': 'SCHOOL_OPS',
}
ALLOWED_STAFF_DEPARTMENTS = {
    'Physics',
    'Mathematics & Statistics',
    'Computer Science & Software Engineering',
    'Senior School Coordinator',
}


def _active_status_to_bool(value):
    return str(value).strip().lower() == 'active'


def _bool_to_active_status(value):
    return 'Active' if value else 'Inactive'


def _serialize_staff(staff):
    return {
        'staff_id': staff.staff_number,
        'first_name': staff.user.first_name,
        'last_name': staff.user.last_name,
        'email': staff.user.email,
        'title': '',
        'department': staff.department.name,
        'active_status': _bool_to_active_status(staff.is_active),
    }


def _serialize_role_assignment_log(log):
    payload = log.changes or {}
    return {
        'id': str(log.log_id),
        'staff_id': payload.get('staff_id', ''),
        'name': payload.get('name', ''),
        'role': payload.get('role', ''),
        'department': payload.get('department', ''),
        'permissions': payload.get('permissions', []),
        'assigned_at': log.created_at.strftime('%Y-%m-%d %H:%M'),
        'status': payload.get('status', 'active'),
    }


def _is_valid_staff_number(value):
    return len(value) == 8 and value.isalnum()


def _build_role_assignments_response():
    logs = (
        AuditLog.objects.filter(changes__kind='ROLE_ASSIGNMENT')
        .exclude(changes__isnull=True)
        .order_by('-created_at')
    )
    items = [_serialize_role_assignment_log(log) for log in logs]
    return Response({'success': True, 'message': 'Role assignments loaded', 'data': {'items': items}})


def _create_role_assignment_response(request):
    payload = request.data or {}
    staff_id = str(payload.get('staff_id', '')).strip()
    role = str(payload.get('role', '')).strip()
    department_name = str(payload.get('department', '')).strip()
    permissions = payload.get('permissions') or []

    errors = {}
    if not _is_valid_staff_number(staff_id):
        errors['staff_id'] = ['staff_id must be 8 characters']
    if role not in ROLE_ALIAS_TO_DB:
        errors['role'] = ['role must be HoD or Admin']
    if not isinstance(permissions, list) or not permissions:
        errors['permissions'] = ['permissions must be a non-empty list']
    if errors:
        return Response(
            {'success': False, 'message': 'Validation failed', 'errors': errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    staff = get_object_or_404(Staff.objects.select_related('user', 'department'), staff_number=staff_id)
    if department_name:
        department, _ = Department.objects.get_or_create(name=department_name)
        staff.department = department
    staff.role = ROLE_ALIAS_TO_DB[role]
    staff.save(update_fields=['role', 'department', 'updated_at'])

    name = staff.user.get_full_name().strip() or staff.user.username
    assignment_log = AuditLog.objects.create(
        action_by=request.staff,
        action_type='COMMENT',
        changes={
            'kind': 'ROLE_ASSIGNMENT',
            'staff_id': staff.staff_number,
            'name': name,
            'role': role,
            'department': staff.department.name,
            'permissions': permissions,
            'status': 'active',
        },
    )

    return Response(
        {
            'success': True,
            'message': 'Role assigned',
            'data': {
                'id': str(assignment_log.log_id),
                'staff_id': staff.staff_number,
                'role': role,
                'department': staff.department.name,
                'status': 'active',
            },
        }
    )


def _read_staff_import_rows(upload_file):
    file_name = (upload_file.name or '').lower()
    if file_name.endswith('.csv'):
        text = upload_file.read().decode('utf-8-sig')
        rows = list(csv.DictReader(io.StringIO(text)))
        return rows, None

    try:
        import openpyxl
    except ImportError:
        return None, 'openpyxl is required for xlsx import'

    if not (file_name.endswith('.xlsx') or file_name.endswith('.xlsm')):
        return None, 'Only .xlsx/.xlsm/.csv files are supported'

    wb = openpyxl.load_workbook(upload_file, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else '' for c in ws[1]]
    rows = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        row_data = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            row_data[header] = row_cells[idx] if idx < len(row_cells) else None
        if not any(v is not None and str(v).strip() != '' for v in row_data.values()):
            continue
        rows.append(row_data)
    return rows, None


def _parse_hos_year_range(request):
    class _ProxyRequest:
        GET = None

    query = request.GET.copy()
    if query.get('from_year') and not query.get('year_from'):
        query['year_from'] = query.get('from_year')
    if query.get('to_year') and not query.get('year_to'):
        query['year_to'] = query.get('to_year')
    proxy = _ProxyRequest()
    proxy.GET = query
    return _parse_year_range(proxy)


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
def hos_staff_list(request):
    """GET /api/headofschool/staff/"""
    qs = Staff.objects.select_related('user', 'department').order_by('staff_number')

    first_name = (request.GET.get('first_name') or '').strip()
    if first_name:
        qs = qs.filter(user__first_name__icontains=first_name)

    last_name = (request.GET.get('last_name') or '').strip()
    if last_name:
        qs = qs.filter(user__last_name__icontains=last_name)

    staff_id = (request.GET.get('staff_id') or '').strip()
    if staff_id:
        qs = qs.filter(staff_number=staff_id)

    try:
        page = max(1, int(request.GET.get('page', 1)))
        page_size = max(1, min(100, int(request.GET.get('page_size', 10))))
    except (TypeError, ValueError):
        return Response(
            {'success': False, 'message': 'page and page_size must be positive integers'},
            status=status.HTTP_400_BAD_REQUEST,
        )

    rows = [_serialize_staff(staff) for staff in qs]
    paginator = Paginator(rows, page_size)
    current_page = paginator.get_page(page)

    return Response(
        {
            'success': True,
            'message': 'Staff loaded',
            'data': {
                'page': current_page.number,
                'page_size': page_size,
                'total': paginator.count,
                'items': list(current_page.object_list),
            },
        }
    )


@api_view(['PATCH'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
@transaction.atomic
def hos_staff_update(request, staff_id):
    """PATCH /api/headofschool/staff/{staff_id}/"""
    payload = request.data or {}
    if str(payload.get('staff_id', '')).strip() != str(staff_id).strip():
        return Response(
            {
                'success': False,
                'message': 'Validation failed',
                'errors': {'staff_id': ['Path staff_id and body staff_id must match']},
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    first_name = str(payload.get('first_name', '')).strip()
    last_name = str(payload.get('last_name', '')).strip()
    email = str(payload.get('email', '')).strip()
    department_name = str(payload.get('department', '')).strip()
    active_status = str(payload.get('active_status', '')).strip()

    errors = {}
    if not _is_valid_staff_number(staff_id):
        errors['staff_id'] = ['staff_id must be 8 characters']
    if not first_name:
        errors['first_name'] = ['first_name is required']
    if not last_name:
        errors['last_name'] = ['last_name is required']
    if '@' not in email:
        errors['email'] = ['email format is invalid']
    if department_name not in ALLOWED_STAFF_DEPARTMENTS:
        errors['department'] = ['department is not in allowed values']
    if active_status not in ('Active', 'Inactive'):
        errors['active_status'] = ['active_status must be Active or Inactive']
    if errors:
        return Response(
            {'success': False, 'message': 'Validation failed', 'errors': errors},
            status=status.HTTP_400_BAD_REQUEST,
        )

    staff = get_object_or_404(Staff.objects.select_related('user', 'department'), staff_number=staff_id)
    department = Department.objects.filter(name=department_name).first()
    if not department:
        department = Department.objects.create(name=department_name)

    staff.user.first_name = first_name
    staff.user.last_name = last_name
    staff.user.email = email
    staff.user.save(update_fields=['first_name', 'last_name', 'email'])

    staff.department = department
    staff.is_active = _active_status_to_bool(active_status)
    staff.save(update_fields=['department', 'is_active', 'updated_at'])

    return Response(
        {
            'success': True,
            'message': 'Staff profile updated',
            'data': {
                'staff_id': staff.staff_number,
                'updated_at': staff.updated_at.isoformat(),
            },
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
def hos_staff_import_template(request):
    """GET /api/headofschool/staff/import-template/"""
    return Response(
        {
            'success': True,
            'message': 'Template ready',
            'data': {
                'file_name': 'Staff_Template.xlsx',
                'download_url': 'http://localhost:8000/media/templates/Staff_Template.xlsx',
            },
        }
    )


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
@transaction.atomic
def hos_staff_import(request):
    """POST /api/headofschool/staff/import/"""
    upload_file = request.FILES.get('file')
    if not upload_file:
        return Response(
            {
                'success': False,
                'message': 'Validation failed',
                'errors': {'file': ['file is required']},
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    rows, error = _read_staff_import_rows(upload_file)
    if error:
        return Response({'success': False, 'message': error}, status=status.HTTP_400_BAD_REQUEST)

    imported_count = 0
    failed_count = 0
    for row in rows:
        staff_number = str(row.get('staff_id', '')).strip()
        first_name = str(row.get('first_name', '')).strip()
        last_name = str(row.get('last_name', '')).strip()
        email = str(row.get('email', '')).strip()
        department_name = str(row.get('department', '')).strip() or 'Computer Science & Software Engineering'
        active_status = str(row.get('active_status', '')).strip() or 'Active'

        if not _is_valid_staff_number(staff_number):
            failed_count += 1
            continue
        if not first_name or not last_name or '@' not in email:
            failed_count += 1
            continue
        if department_name not in ALLOWED_STAFF_DEPARTMENTS:
            failed_count += 1
            continue
        if active_status not in ('Active', 'Inactive'):
            failed_count += 1
            continue

        department, _ = Department.objects.get_or_create(name=department_name)
        user, created_user = User.objects.get_or_create(
            username=staff_number,
            defaults={
                'first_name': first_name,
                'last_name': last_name,
                'email': email,
            },
        )
        if created_user:
            user.set_unusable_password()
            user.save()
        else:
            user.first_name = first_name
            user.last_name = last_name
            user.email = email
            user.save(update_fields=['first_name', 'last_name', 'email'])

        Staff.objects.update_or_create(
            staff_number=staff_number,
            defaults={
                'user': user,
                'role': 'ACADEMIC',
                'department': department,
                'is_active': _active_status_to_bool(active_status),
            },
        )
        imported_count += 1

    return Response(
        {
            'success': True,
            'message': 'Staff import completed',
            'data': {
                'imported_count': imported_count,
                'failed_count': failed_count,
            },
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
def hos_role_assignments(request):
    """GET /api/headofschool/role-assignments/"""
    return _build_role_assignments_response()


@api_view(['GET', 'POST'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
def hos_role_assignments_collection(request):
    if request.method == 'GET':
        return _build_role_assignments_response()
    return _create_role_assignment_response(request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
@transaction.atomic
def hos_create_role_assignment(request):
    """POST /api/headofschool/role-assignments/"""
    return _create_role_assignment_response(request)


@api_view(['POST'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
@transaction.atomic
def hos_disable_role_assignment(request, id):
    """POST /api/headofschool/role-assignments/{id}/disable/"""
    reason = str((request.data or {}).get('reason', '')).strip()
    source_log = get_object_or_404(AuditLog, log_id=id, changes__kind='ROLE_ASSIGNMENT')
    payload = source_log.changes or {}
    staff_id = payload.get('staff_id')
    staff = get_object_or_404(Staff, staff_number=staff_id)
    staff.role = 'ACADEMIC'
    staff.save(update_fields=['role', 'updated_at'])

    disabled_payload = dict(payload)
    disabled_payload['status'] = 'disabled'
    disabled_payload['kind'] = 'ROLE_ASSIGNMENT'
    disabled_payload['disabled_from'] = str(source_log.log_id)
    disabled_payload['reason'] = reason
    disabled_log = AuditLog.objects.create(
        action_by=request.staff,
        action_type='COMMENT',
        changes=disabled_payload,
        comment=reason or None,
    )

    return Response(
        {
            'success': True,
            'message': 'Role assignment disabled',
            'data': {
                'id': str(disabled_log.log_id),
                'status': 'disabled',
            },
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
def hos_visualization(request):
    """GET /api/headofschool/visualization/"""
    year_from, year_to = _parse_hos_year_range(request)
    semester = request.GET.get('semester', 'All')
    department_filter = (request.GET.get('department') or 'All Departments').strip()

    confirmed_subq = AuditLog.objects.filter(
        report=OuterRef('pk'),
        changes__kind='CONFIRMATION',
        changes__confirmation='confirmed',
    )
    base_qs = (
        WorkloadReport.objects.filter(is_current=True)
        .select_related('snapshot_department')
        .prefetch_related('items')
        .annotate(is_confirmed=Exists(confirmed_subq))
        .filter(Q(status__in=['PENDING', 'APPROVED', 'REJECTED']) | Q(status='INITIAL', is_confirmed=True))
    )
    base_qs = _filter_reports_by_range(base_qs, year_from, year_to, semester)
    if department_filter and department_filter != 'All Departments':
        base_qs = base_qs.filter(snapshot_department__name=department_filter)

    reports = list(base_qs.order_by('academic_year', 'semester', 'snapshot_department__name'))

    summary = {
        'total_departments': len({report.snapshot_department_id for report in reports}),
        'total_academics': len({report.staff_id for report in reports}),
        'total_work_hours': float(
            round(sum(sum(item.allocated_hours for item in report.items.all()) for report in reports), 2)
        ),
        'pending_requests': sum(1 for report in reports if report.status == 'PENDING'),
        'approved_requests': sum(1 for report in reports if report.status == 'APPROVED'),
        'rejected_requests': sum(1 for report in reports if report.status == 'REJECTED'),
    }

    department_stats_map = {}
    for report in reports:
        key = report.snapshot_department.name
        stats = department_stats_map.setdefault(
            key,
            {
                'department': key,
                'academics': set(),
                'total_hours': Decimal('0.00'),
                'pending': 0,
                'approved': 0,
                'rejected': 0,
            },
        )
        stats['academics'].add(report.staff_id)
        stats['total_hours'] += sum(item.allocated_hours for item in report.items.all())
        if report.status == 'PENDING':
            stats['pending'] += 1
        elif report.status == 'APPROVED':
            stats['approved'] += 1
        elif report.status == 'REJECTED':
            stats['rejected'] += 1

    department_stats = []
    for stats in department_stats_map.values():
        department_stats.append(
            {
                'department': stats['department'],
                'academics': len(stats['academics']),
                'total_hours': float(round(stats['total_hours'], 2)),
                'pending': stats['pending'],
                'approved': stats['approved'],
                'rejected': stats['rejected'],
            }
        )
    department_stats.sort(key=lambda item: item['department'])

    trend_map = {}
    for report in reports:
        semester_label = f"{report.academic_year} {report.semester}"
        trend_row = trend_map.setdefault(semester_label, {'semester': semester_label})
        trend_row[report.snapshot_department.name] = float(
            round(
                trend_row.get(report.snapshot_department.name, 0)
                + sum(item.allocated_hours for item in report.items.all()),
                2,
            )
        )
    workload_trend = [trend_map[key] for key in sorted(trend_map.keys())]

    return Response(
        {
            'success': True,
            'message': 'Visualization loaded',
            'data': {
                'reporting_period_label': _reporting_period_label(year_from, year_to, semester),
                'scope_label': department_filter or 'All Departments',
                'summary': summary,
                'department_stats': department_stats,
                'workload_trend': workload_trend,
            },
        }
    )


@api_view(['GET'])
@permission_classes([IsAuthenticated])
@require_role('HOS', 'SCHOOL_OPS')
def hos_export(request):
    """GET /api/headofschool/export/"""
    try:
        import openpyxl
    except ImportError:
        return Response(
            {'success': False, 'message': 'Export unavailable: openpyxl not installed'},
            status=status.HTTP_503_SERVICE_UNAVAILABLE,
        )

    year_from, year_to = _parse_hos_year_range(request)
    semester = request.GET.get('semester', 'All')
    department_filter = (request.GET.get('department') or 'All Departments').strip()

    qs = (
        WorkloadReport.objects.filter(is_current=True)
        .select_related('staff__user', 'snapshot_department')
        .prefetch_related('items')
    )
    qs = _filter_reports_by_range(qs, year_from, year_to, semester).order_by(
        'snapshot_department__name', 'staff__staff_number', 'academic_year', 'semester'
    )
    if department_filter and department_filter != 'All Departments':
        qs = qs.filter(snapshot_department__name=department_filter)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'HeadOfSchool Export'
    ws.append(
        [
            'Staff Number',
            'Name',
            'Department',
            'Academic Year',
            'Semester',
            'Category',
            'Unit Code',
            'Description',
            'Hours',
            'Status',
            'Export Date',
        ]
    )

    export_date = date.today().isoformat()
    for report in qs:
        user = report.staff.user
        full_name = user.get_full_name().strip() or user.username
        items = list(report.items.all())
        if not items:
            ws.append(
                [
                    report.staff.staff_number,
                    full_name,
                    report.snapshot_department.name,
                    report.academic_year,
                    report.semester,
                    '',
                    '',
                    '',
                    0,
                    report.status.lower(),
                    export_date,
                ]
            )
        else:
            for item in items:
                ws.append(
                    [
                        report.staff.staff_number,
                        full_name,
                        report.snapshot_department.name,
                        report.academic_year,
                        report.semester,
                        item.category,
                        item.unit_code or '',
                        item.description or '',
                        float(item.allocated_hours),
                        report.status.lower(),
                        export_date,
                    ]
                )

    file_name = f'HeadOfSchool_Workload_{export_date}.xlsx'
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{file_name}"'
    return response
